"""Platform-admin API — Task 6.1.

Every route on this router is admin-only: the router carries a router-level
dependency that 403s any caller whose role is not ``admin`` (fail-closed —
adding a route here can never accidentally ship unguarded). Every mutation
writes an audit row via ``server.write_audit``.

Groups (see docs/superpowers/audits/2026-07-07-admin-api-audit.md §CONSOLIDATED):
users, organizations, master-data, terms, tickets, audit, alerts,
notification-monitoring, reports, delegations, emergency (break-the-glass),
invitations, messages (broadcasts), admin trials.

Patient PII rules:
- user lists / exports pseudonymize patient names + contact details
- trial reads return AGGREGATES + masked subjects (SUBJ-xxx + initials) unless
  the caller holds an ACTIVE break-the-glass session — and every unmasked read
  during a session is itself audited with the session id.
"""
from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import logging
import os
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel, EmailStr, Field
from pymongo import ReturnDocument
from starlette.concurrency import run_in_threadpool

import otp_service
import storage as file_storage
from server import (
    INVITE_TTL_DAYS,
    ORG_TYPES,
    Role,
    _invitation_status,
    _invite_link,
    new_invite_code,
    _parse_ymd,
    current_user,
    db,
    iso,
    now,
    pwd_ctx,
    require_roles,
    serialize,
    write_audit,
)

# Router-level guard: EVERY admin route 403s non-admin callers.
router = APIRouter(prefix='/api/admin', dependencies=[Depends(require_roles('admin'))])

USER_PROJECTION = {'_id': 0, 'hashed_password': 0, 'security_answer_hash': 0}
PASSWORD_RESET_TTL_MIN = 30


# ── PII masking helpers ──────────────────────────────────────────────────────
def _mask_name(name: str) -> str:
    parts = [p for p in (name or '').split() if p]
    return ' '.join(f'{p[0].upper()}***' for p in parts) or 'U***'


def _mask_email(email: str) -> str:
    email = email or ''
    if '@' not in email:
        return '***' if email else ''
    local, dom = email.split('@', 1)
    return f'{local[:1]}***@{dom}'


def _mask_phone(phone: str) -> str:
    p = (phone or '').strip()
    if not p:
        return ''
    if len(p) < 6:
        return '***'
    return p[:3] + '*' * (len(p) - 5) + p[-2:]


def _pseudonymize_patient(u: dict) -> dict:
    """Patients are pseudonymized in admin lists/exports (regulated app)."""
    if u.get('role') == 'patient':
        u = dict(u)
        u['full_name'] = _mask_name(u.get('full_name', ''))
        u['email'] = _mask_email(u.get('email', ''))
        u['phone'] = _mask_phone(u.get('phone', ''))
        u['pseudonymized'] = True
    return u


def _masked_subject(p: dict) -> dict:
    """Trial subject rows carry NO PII: stable pseudo-label + initials only."""
    return {
        'subject': f"SUBJ-{(p.get('id') or '')[-3:]}",
        'initials': p.get('avatar_initials', ''),
        'status': p.get('status', ''),
        'enrolled_date': p.get('enrolled_date', ''),
    }


def _user_status(u: dict) -> str:
    if u.get('status'):
        return u['status']
    if u.get('lock_info'):
        return 'Locked'
    return 'Active'


async def _find_or_404(coll, doc_id: str, what: str) -> dict:
    doc = await coll.find_one({'id': doc_id}, {'_id': 0})
    if not doc:
        raise HTTPException(404, f'{what} not found')
    return doc


async def _deliver_password_reset_link(user: dict, admin: dict, purpose: str) -> dict:
    """Create and email an expiring single-use token. The raw token is never
    stored, logged, audited, or returned to the platform-admin client."""
    token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(token.encode('utf-8')).hexdigest()
    issued_at = now()
    expires_at = issued_at + timedelta(minutes=PASSWORD_RESET_TTL_MIN)
    base = os.environ.get('PASSWORD_RESET_URL', 'mytrialboard://reset-password')
    separator = '&' if '?' in base else '?'
    reset_link = f'{base}{separator}token={token}'
    doc = {
        'id': str(uuid.uuid4()),
        'token_hash': token_hash,
        'user_id': user['id'],
        'email': user.get('email', ''),
        'purpose': purpose,
        'requested_by': admin['id'],
        'created_at': issued_at,
        'expires_at': expires_at,
        'used_at': None,
        'revoked_at': None,
    }
    await db.password_reset_tokens.update_many(
        {'user_id': user['id'], 'used_at': None, 'revoked_at': None},
        {'$set': {'revoked_at': issued_at}})
    await db.password_reset_tokens.insert_one(doc)
    try:
        await run_in_threadpool(
            otp_service.send_password_reset_email,
            user['email'],
            reset_link,
            PASSWORD_RESET_TTL_MIN,
        )
    except otp_service.OTPConfigError:
        await db.password_reset_tokens.delete_one({'id': doc['id']})
        raise HTTPException(503, 'Password-reset email delivery is not configured.')
    except otp_service.OTPDeliveryError:
        await db.password_reset_tokens.delete_one({'id': doc['id']})
        raise HTTPException(502, 'The password-reset email could not be delivered.')
    return {
        'reset_sent': True,
        'delivery_channel': 'email',
        'email': _mask_email(user.get('email', '')),
        'expires_at': iso(expires_at),
    }


# ═════════════════════════════════════════════════════════════════════════════
# USERS
# ═════════════════════════════════════════════════════════════════════════════
class AdminUserCreate(BaseModel):
    email: EmailStr
    full_name: str = Field(min_length=1)
    role: Role
    phone: Optional[str] = ''
    organization: Optional[str] = ''
    send_invite: bool = True


class UserStatusIn(BaseModel):
    status: Literal['Active', 'Suspended']
    reason: Optional[str] = ''


class UnlockIn(BaseModel):
    identity_checks: List[str] = Field(min_length=2)
    reason: str = Field(min_length=10)
    force_password_reset: bool = False


@router.get('/users')
async def admin_list_users(search: Optional[str] = None, role: Optional[str] = None,
                           status: Optional[str] = None, limit: int = Query(500, le=2000)):
    q: Dict = {}
    if role:
        q['role'] = role
    if search:
        rx = {'$regex': search, '$options': 'i'}
        q['$or'] = [{'full_name': rx}, {'email': rx}, {'organization': rx}]
    rows = await db.users.find(q, USER_PROJECTION).sort('created_at', -1).to_list(limit)
    out = []
    for u in rows:
        u['status'] = _user_status(u)
        if status and u['status'] != status:
            continue
        out.append(_pseudonymize_patient(u))
    return out


@router.get('/users/export')
async def admin_export_users(admin=Depends(current_user)):
    """CSV export of the user directory (patients pseudonymized)."""
    rows = await db.users.find({}, USER_PROJECTION).sort('created_at', -1).to_list(5000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['id', 'name', 'email', 'phone', 'role', 'organization', 'status', 'created_at'])
    for u in rows:
        u = _pseudonymize_patient(u)
        w.writerow([u.get('id'), u.get('full_name'), u.get('email'), u.get('phone'),
                    u.get('role'), u.get('organization'), _user_status(u), iso(u.get('created_at'))])
    await write_audit(admin, 'admin.users_export', f'Exported {len(rows)} users to CSV')
    return Response(content=buf.getvalue(), media_type='text/csv',
                    headers={'Content-Disposition': 'attachment; filename="users.csv"'})


@router.post('/users')
async def admin_create_user(body: AdminUserCreate, admin=Depends(current_user)):
    email = body.email.lower()
    if await db.users.find_one({'email': email}):
        raise HTTPException(400, 'Email already registered')
    inaccessible_password = secrets.token_urlsafe(48)
    doc = {
        'id': str(uuid.uuid4()), 'email': email, 'full_name': body.full_name.strip(),
        'role': body.role, 'phone': body.phone or '', 'organization': body.organization or '',
        'hashed_password': pwd_ctx.hash(inaccessible_password),
        'security_question': '', 'security_answer_hash': '',
        'avatar_initials': ''.join(w[0].upper() for w in body.full_name.split()[:2]) or 'U',
        'status': 'Pending Verification', 'must_reset_password': True,
        'created_at': now(), 'is_online': False, 'created_by_admin': admin['id'],
    }
    await db.users.insert_one(doc)
    invitation = None
    if body.send_invite:
        inv = {
            'id': str(uuid.uuid4()), 'token': new_invite_code(), 'email': email,
            'phone': body.phone or '', 'full_name': body.full_name, 'role': body.role,
            'trial_id': None, 'invited_by': admin['id'],
            'org': (body.organization or '').strip(), 'site': '',
            'inviter_name': admin.get('full_name') or '',
            'inviter_organization': (
                body.organization or admin.get('organization') or 'My Trial Board').strip(),
            'status': 'pending', 'created_at': now(),
            'expires_at': now() + timedelta(days=INVITE_TTL_DAYS), 'resend_count': 0,
        }
        await db.invitations.insert_one(inv)
        try:
            await run_in_threadpool(
                otp_service.send_invitation_email,
                email,
                _invite_link(inv['token']),
                inv['full_name'],
                inv['inviter_name'],
                inv['inviter_organization'],
            )
        except (otp_service.OTPConfigError, otp_service.OTPDeliveryError):
            await db.invitations.delete_one({'id': inv['id']})
            await db.users.delete_one({'id': doc['id']})
            raise HTTPException(502, 'The invitation email could not be delivered.')
        invitation = {**serialize(inv), 'invite_link': _invite_link(inv['token'])}
    try:
        password_setup = await _deliver_password_reset_link(doc, admin, 'account_setup')
    except HTTPException:
        await db.users.delete_one({'id': doc['id']})
        if invitation:
            await db.invitations.delete_one({'id': invitation['id']})
        raise
    await write_audit(admin, 'admin.user_create',
                      f"Created user {email} ({body.role}) and sent a password setup link",
                      target_id=doc['id'], password_setup_sent=True)
    return {'user': serialize({**doc}), 'invitation': invitation,
            'password_setup': password_setup}


@router.get('/users/{user_id}')
async def admin_get_user(user_id: str):
    u = await db.users.find_one({'id': user_id}, USER_PROJECTION)
    if not u:
        raise HTTPException(404, 'User not found')
    u['status'] = _user_status(u)
    return _pseudonymize_patient(u)


@router.patch('/users/{user_id}/status')
async def admin_set_user_status(user_id: str, body: UserStatusIn, admin=Depends(current_user)):
    u = await _find_or_404(db.users, user_id, 'User')
    updates: Dict = {'status': body.status}
    if body.status == 'Suspended':
        updates['force_logout_at'] = now()   # kill active sessions immediately
        await db.refresh_tokens.update_many(
            {'user_id': user_id, 'status': 'active'},
            {'$set': {'status': 'revoked', 'revoked_at': updates['force_logout_at'],
                      'revoke_reason': 'account suspended'}},
        )
    await db.users.update_one({'id': user_id}, {'$set': updates})
    await write_audit(admin, 'admin.user_status',
                      f"Set {u.get('email')} status to {body.status}"
                      + (f" — {body.reason}" if body.reason else ''),
                      target_id=user_id)
    return {'ok': True, 'id': user_id, 'status': body.status}


@router.post('/users/{user_id}/unlock')
async def admin_unlock_user(user_id: str, body: UnlockIn, admin=Depends(current_user)):
    """Unlock requires ≥2 completed identity checks and a reason ≥10 chars
    (regulatory traceability). Optionally forces a password reset."""
    u = await _find_or_404(db.users, user_id, 'User')
    checks = [c.strip() for c in body.identity_checks if c and c.strip()]
    if len(checks) < 2:
        raise HTTPException(400, 'At least 2 identity checks are required to unlock')
    updates: Dict = {'status': 'Active', 'failed_attempts': 0}
    if body.force_password_reset:
        updates['must_reset_password'] = True
    await db.users.update_one({'id': user_id},
                              {'$set': updates, '$unset': {'lock_info': ''}})
    await write_audit(admin, 'admin.user_unlock',
                      f"Unlocked {u.get('email')} — {body.reason}",
                      target_id=user_id, identity_checks=checks,
                      force_password_reset=body.force_password_reset)
    return {'ok': True, 'id': user_id, 'status': 'Active'}


@router.post('/users/{user_id}/reset-password')
async def admin_reset_password(user_id: str, admin=Depends(current_user)):
    u = await _find_or_404(db.users, user_id, 'User')
    delivery = await _deliver_password_reset_link(u, admin, 'password_reset')
    await db.users.update_one({'id': user_id}, {'$set': {
        'must_reset_password': True, 'force_logout_at': now(), 'is_online': False}})
    await db.refresh_tokens.update_many(
        {'user_id': user_id, 'status': 'active'},
        {'$set': {'status': 'revoked', 'revoked_at': now(),
                  'revoke_reason': 'password reset requested'}},
    )
    await write_audit(admin, 'admin.user_reset_password',
                      f"Sent password reset link for {u.get('email')}",
                      target_id=user_id, delivery_channel='email')
    return {'ok': True, 'id': user_id, **delivery}


@router.post('/users/{user_id}/force-logout')
async def admin_force_logout(user_id: str, admin=Depends(current_user)):
    u = await _find_or_404(db.users, user_id, 'User')
    revoked_at = now()
    await db.users.update_one({'id': user_id}, {'$set': {
        'force_logout_at': revoked_at, 'is_online': False}})
    await db.refresh_tokens.update_many(
        {'user_id': user_id, 'status': 'active'},
        {'$set': {'status': 'revoked', 'revoked_at': revoked_at,
                  'revoke_reason': 'administrative force logout'}},
    )
    await write_audit(admin, 'admin.user_force_logout',
                      f"Forced logout for {u.get('email')}", target_id=user_id)
    return {'ok': True, 'id': user_id}


# ═════════════════════════════════════════════════════════════════════════════
# ORGANIZATIONS
# ═════════════════════════════════════════════════════════════════════════════
class OrgCreate(BaseModel):
    name: str = Field(min_length=1)
    type: Literal['sponsor', 'cro', 'smo', 'site']
    address: Optional[str] = ''
    contact: Optional[str] = ''
    email: Optional[str] = ''
    website: Optional[str] = ''


class OrgPatch(BaseModel):
    name: Optional[str] = None
    type: Optional[Literal['sponsor', 'cro', 'smo', 'site']] = None
    address: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    status: Optional[Literal['active', 'suspended']] = None


class OrgMergeIn(BaseModel):
    target_org_id: str
    justification: str = Field(min_length=10)


class NameRequestApproveIn(BaseModel):
    finalName: str = Field(min_length=1)


class NameRequestRejectIn(BaseModel):
    reason: str = Field(min_length=1)


async def _org_counts(org: dict) -> dict:
    org = dict(org)
    org['users'] = await db.users.count_documents({'organization': org['name']})
    org['trials'] = await db.trials.count_documents({'sponsor_name': org['name']})
    return org


@router.get('/organizations')
async def admin_list_orgs(type: Optional[str] = None, search: Optional[str] = None,
                          status: Optional[str] = None):
    q: Dict = {}
    if type:
        q['type'] = type
    if status:
        q['status'] = status
    if search:
        q['name'] = {'$regex': search, '$options': 'i'}
    rows = await db.organizations.find(q, {'_id': 0}).sort('name', 1).to_list(1000)
    return [await _org_counts(o) for o in rows]


@router.get('/organizations/duplicates')
async def admin_org_duplicates():
    """Groups of organizations whose normalized names collide (merge candidates)."""
    rows = await db.organizations.find({'status': {'$ne': 'merged'}}, {'_id': 0}).to_list(2000)
    groups: Dict[str, List[dict]] = {}
    for o in rows:
        key = ''.join(ch for ch in (o.get('name') or '').lower() if ch.isalnum())
        groups.setdefault(key, []).append(o)
    return [{'key': k, 'organizations': v} for k, v in groups.items() if len(v) > 1]


@router.get('/organizations/name-requests')
async def admin_org_name_requests(status: Optional[str] = None):
    q: Dict = {}
    if status:
        q['status'] = status
    return await db.org_name_requests.find(q, {'_id': 0}).sort('created_at', -1).to_list(500)


@router.post('/organizations/name-requests/{request_id}/approve')
async def admin_approve_name_request(request_id: str, body: NameRequestApproveIn,
                                     admin=Depends(current_user)):
    req = await _find_or_404(db.org_name_requests, request_id, 'Name-correction request')
    if req.get('status') != 'pending':
        raise HTTPException(400, 'This request has already been actioned')
    org = await db.organizations.find_one({'id': req.get('org_id')}, {'_id': 0})
    final_name = body.finalName.strip()
    if org:
        old_name = org['name']
        await db.organizations.update_one({'id': org['id']}, {'$set': {'name': final_name}})
        # Keep membership consistent: users carry the org NAME string.
        await db.users.update_many({'organization': old_name},
                                   {'$set': {'organization': final_name}})
    await db.org_name_requests.update_one({'id': request_id}, {'$set': {
        'status': 'approved', 'finalName': final_name,
        'actioned_by': admin['full_name'], 'actioned_at': now()}})
    await write_audit(admin, 'admin.org_name_approve',
                      f"Approved org name correction → \"{final_name}\"",
                      target_id=request_id, org_id=req.get('org_id'))
    return {'ok': True, 'id': request_id, 'finalName': final_name}


@router.post('/organizations/name-requests/{request_id}/reject')
async def admin_reject_name_request(request_id: str, body: NameRequestRejectIn,
                                    admin=Depends(current_user)):
    req = await _find_or_404(db.org_name_requests, request_id, 'Name-correction request')
    if req.get('status') != 'pending':
        raise HTTPException(400, 'This request has already been actioned')
    await db.org_name_requests.update_one({'id': request_id}, {'$set': {
        'status': 'rejected', 'rejectReason': body.reason,
        'actioned_by': admin['full_name'], 'actioned_at': now()}})
    await write_audit(admin, 'admin.org_name_reject',
                      f"Rejected org name correction — {body.reason}", target_id=request_id)
    return {'ok': True, 'id': request_id, 'status': 'rejected'}


@router.post('/organizations')
async def admin_create_org(body: OrgCreate, admin=Depends(current_user)):
    name = body.name.strip()
    if await db.organizations.find_one({'name': name}):
        raise HTTPException(400, 'An organization with this name already exists')
    doc = {
        'id': str(uuid.uuid4()), 'name': name, 'type': body.type,
        'address': body.address or '', 'contact': body.contact or '',
        'email': body.email or '', 'website': body.website or '',
        'status': 'active', 'created_at': now(), 'created_by': admin['id'],
    }
    await db.organizations.insert_one(doc)
    await write_audit(admin, 'admin.org_create', f'Created organization "{name}"',
                      target_id=doc['id'])
    return serialize({**doc})


@router.patch('/organizations/{org_id}')
async def admin_patch_org(org_id: str, body: OrgPatch, admin=Depends(current_user)):
    org = await _find_or_404(db.organizations, org_id, 'Organization')
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not updates:
        return org
    new_name = updates.get('name', '').strip() if 'name' in updates else None
    if new_name and new_name != org['name']:
        if await db.organizations.find_one({'name': new_name, 'id': {'$ne': org_id}}):
            raise HTTPException(400, 'An organization with this name already exists')
        updates['name'] = new_name
        await db.users.update_many({'organization': org['name']},
                                   {'$set': {'organization': new_name}})
    await db.organizations.update_one({'id': org_id}, {'$set': updates})
    await write_audit(admin, 'admin.org_update',
                      f"Updated organization \"{org['name']}\" ({', '.join(updates)})",
                      target_id=org_id, changes=updates)
    return await db.organizations.find_one({'id': org_id}, {'_id': 0})


@router.post('/organizations/{org_id}/merge')
async def admin_merge_orgs(org_id: str, body: OrgMergeIn, admin=Depends(current_user)):
    """Merge org {org_id} INTO target_org_id. Irreversible: users and trials are
    repointed to the target; the source is tombstoned as status='merged'."""
    source = await _find_or_404(db.organizations, org_id, 'Organization')
    target = await _find_or_404(db.organizations, body.target_org_id, 'Target organization')
    if source['id'] == target['id']:
        raise HTTPException(400, 'Cannot merge an organization into itself')
    if source.get('status') == 'merged':
        raise HTTPException(400, 'This organization has already been merged')
    moved_users = await db.users.update_many(
        {'organization': source['name']}, {'$set': {'organization': target['name']}})
    moved_trials = await db.trials.update_many(
        {'sponsor_name': source['name']}, {'$set': {'sponsor_name': target['name']}})
    await db.organizations.update_one({'id': source['id']}, {'$set': {
        'status': 'merged', 'merged_into': target['id'], 'merged_at': now(),
        'merge_justification': body.justification}})
    await write_audit(admin, 'admin.org_merge',
                      f"Merged \"{source['name']}\" into \"{target['name']}\" — {body.justification}",
                      target_id=source['id'], merged_into=target['id'],
                      moved_users=moved_users.modified_count,
                      moved_trials=moved_trials.modified_count)
    return {'ok': True, 'merged': source['id'], 'into': target['id'],
            'moved_users': moved_users.modified_count,
            'moved_trials': moved_trials.modified_count}


# ═════════════════════════════════════════════════════════════════════════════
# MASTER DATA ("Others: specify" queue + global values)
# ═════════════════════════════════════════════════════════════════════════════
class MasterDataApproveIn(BaseModel):
    value: Optional[str] = None   # edit-and-approve when provided


class MasterDataRejectIn(BaseModel):
    reason: str = Field(min_length=1)


@router.get('/master-data/submissions')
async def admin_master_data_submissions(status: Optional[str] = None,
                                        fieldType: Optional[str] = None):
    q: Dict = {}
    if status:
        q['status'] = status
    if fieldType:
        q['fieldType'] = fieldType
    return await db.master_data_submissions.find(q, {'_id': 0}) \
        .sort('dateSubmitted', -1).to_list(500)


@router.post('/master-data/submissions/{submission_id}/approve')
async def admin_approve_master_data(submission_id: str, body: MasterDataApproveIn,
                                    admin=Depends(current_user)):
    sub = await _find_or_404(db.master_data_submissions, submission_id, 'Submission')
    if sub.get('status') != 'pending':
        raise HTTPException(400, 'This submission has already been actioned')
    final_value = (body.value or sub['value']).strip()
    await db.master_data_submissions.update_one({'id': submission_id}, {'$set': {
        'status': 'approved', 'value': final_value,
        'actionBy': admin['full_name'], 'actioned_at': now()}})
    await db.master_data_values.update_one(
        {'fieldType': sub['fieldType'], 'value': final_value},
        {'$setOnInsert': {'id': str(uuid.uuid4()), 'added_by': admin['full_name'],
                          'added_at': now(), 'source_submission': submission_id}},
        upsert=True)
    if sub.get('submittedById') and sub.get('fieldType') == 'department':
        await db.users.update_one({'id': sub['submittedById']}, {'$set': {
            'profile.department': final_value,
            'profile.department_is_custom': False,
            'profile.department_review_status': 'approved',
        }})
    await write_audit(admin, 'admin.master_data_approve',
                      f"Approved {sub['fieldType']} value \"{final_value}\""
                      + (' (edited)' if body.value else ''),
                      target_id=submission_id)
    return {'ok': True, 'id': submission_id, 'status': 'approved', 'value': final_value}


@router.post('/master-data/submissions/{submission_id}/reject')
async def admin_reject_master_data(submission_id: str, body: MasterDataRejectIn,
                                   admin=Depends(current_user)):
    sub = await _find_or_404(db.master_data_submissions, submission_id, 'Submission')
    if sub.get('status') != 'pending':
        raise HTTPException(400, 'This submission has already been actioned')
    await db.master_data_submissions.update_one({'id': submission_id}, {'$set': {
        'status': 'rejected', 'rejectReason': body.reason,
        'actionBy': admin['full_name'], 'actioned_at': now()}})
    if sub.get('submittedById') and sub.get('fieldType') == 'department':
        await db.users.update_one({'id': sub['submittedById']}, {'$set': {
            'profile.department_is_custom': True,
            'profile.department_review_status': 'rejected',
        }})
    await write_audit(admin, 'admin.master_data_reject',
                      f"Rejected {sub['fieldType']} value \"{sub['value']}\" — {body.reason}",
                      target_id=submission_id)
    return {'ok': True, 'id': submission_id, 'status': 'rejected'}


@router.get('/master-data/values')
async def admin_master_data_values(fieldType: Optional[str] = None):
    q: Dict = {}
    if fieldType:
        q['fieldType'] = fieldType
    return await db.master_data_values.find(q, {'_id': 0}).sort('value', 1).to_list(1000)


# ═════════════════════════════════════════════════════════════════════════════
# INVITATIONS (admin-wide — no org restriction, unlike the staff endpoints)
# ═════════════════════════════════════════════════════════════════════════════
class AdminInvitationIn(BaseModel):
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    full_name: Optional[str] = ''
    designation: Optional[str] = ''
    role: Role = 'patient'
    entityType: Optional[str] = ''
    organization: Optional[str] = ''
    site: Optional[str] = ''
    trial_id: Optional[str] = None


@router.get('/invitations')
async def admin_list_invitations(status: Optional[str] = None):
    rows = await db.invitations.find({}, {'_id': 0}).sort('created_at', -1).to_list(1000)
    out = []
    for inv in rows:
        inv['status'] = _invitation_status(inv)
        if status and inv['status'] != status:
            continue
        out.append(inv)
    return out


@router.post('/invitations')
async def admin_create_invitation(body: AdminInvitationIn, admin=Depends(current_user)):
    if not body.email and not body.phone:
        raise HTTPException(400, 'Email or phone required')
    token = new_invite_code()
    doc = {
        'id': str(uuid.uuid4()), 'token': token,
        'email': (body.email or '').lower(), 'phone': body.phone or '',
        'full_name': body.full_name or '', 'designation': body.designation or '',
        'role': body.role, 'entityType': body.entityType or '',
        'trial_id': body.trial_id, 'invited_by': admin['id'],
        'org': (body.organization or '').strip(), 'site': (body.site or '').strip(),
        'inviter_name': admin.get('full_name') or '',
        'inviter_organization': (
            body.organization or admin.get('organization') or 'My Trial Board').strip(),
        'status': 'pending', 'created_at': now(),
        'expires_at': now() + timedelta(days=INVITE_TTL_DAYS), 'resend_count': 0,
    }
    await db.invitations.insert_one(doc)
    if doc['email']:
        try:
            await run_in_threadpool(
                otp_service.send_invitation_email,
                doc['email'],
                _invite_link(doc['token']),
                doc['full_name'],
                doc['inviter_name'],
                doc['inviter_organization'],
            )
        except (otp_service.OTPConfigError, otp_service.OTPDeliveryError):
            await db.invitations.delete_one({'id': doc['id']})
            raise HTTPException(502, 'The invitation email could not be delivered.')
    await write_audit(admin, 'admin.invitation_create',
                      f"Invited {doc['email'] or doc['phone']} as {doc['role']}",
                      target_id=doc['id'])
    return {**serialize(doc), 'invite_link': _invite_link(token)}


@router.post('/invitations/{invitation_id}/resend')
async def admin_resend_invitation(invitation_id: str, admin=Depends(current_user)):
    inv = await _find_or_404(db.invitations, invitation_id, 'Invitation')
    if _invitation_status(inv) not in ('pending', 'expired'):
        raise HTTPException(400, 'Only pending or expired invitations can be resent')
    new_exp = now() + timedelta(days=INVITE_TTL_DAYS)
    await db.invitations.update_one({'id': invitation_id}, {
        '$set': {'status': 'pending', 'expires_at': new_exp, 'last_sent_at': now()},
        '$inc': {'resend_count': 1}})
    if inv.get('email'):
        original_inviter = await db.users.find_one(
            {'id': inv.get('invited_by')},
            {'_id': 0, 'full_name': 1, 'organization': 1},
        ) or {}
        try:
            await run_in_threadpool(
                otp_service.send_invitation_email,
                inv['email'],
                _invite_link(inv['token']),
                inv.get('full_name', ''),
                inv.get('inviter_name') or original_inviter.get('full_name')
                or admin.get('full_name') or '',
                inv.get('inviter_organization')
                or original_inviter.get('organization') or inv.get('org')
                or 'My Trial Board',
            )
        except (otp_service.OTPConfigError, otp_service.OTPDeliveryError):
            raise HTTPException(502, 'The invitation email could not be delivered.')
    await write_audit(admin, 'admin.invitation_resend',
                      f"Resent invitation for {inv.get('email') or inv.get('phone')}",
                      target_id=invitation_id)
    return {'ok': True, 'invite_link': _invite_link(inv['token']), 'expires_at': iso(new_exp)}


@router.post('/invitations/{invitation_id}/cancel')
async def admin_cancel_invitation(invitation_id: str, admin=Depends(current_user)):
    inv = await _find_or_404(db.invitations, invitation_id, 'Invitation')
    if inv.get('status') == 'accepted':
        raise HTTPException(400, 'An accepted invitation cannot be cancelled')
    await db.invitations.update_one({'id': invitation_id},
                                    {'$set': {'status': 'cancelled', 'cancelled_at': now()}})
    await write_audit(admin, 'admin.invitation_cancel',
                      f"Cancelled invitation for {inv.get('email') or inv.get('phone')}",
                      target_id=invitation_id)
    return {'ok': True, 'status': 'cancelled'}


# ═════════════════════════════════════════════════════════════════════════════
# SUPPORT TICKETS (admin triage — the user-side endpoints stay untouched)
# ═════════════════════════════════════════════════════════════════════════════
class TicketNoteIn(BaseModel):
    text: str = Field(min_length=1)


class TicketPatch(BaseModel):
    status: Optional[Literal['Open', 'In Progress', 'Resolved', 'Closed']] = None
    priority: Optional[Literal['low', 'medium', 'high', 'urgent']] = None


async def _enrich_ticket(t: dict) -> dict:
    t = dict(t)
    u = await db.users.find_one({'id': t.get('user_id')}, USER_PROJECTION)
    if u:
        u = _pseudonymize_patient(u)
        t['user'] = {'id': u['id'], 'name': u.get('full_name', ''),
                     'email': u.get('email', ''), 'role': u.get('role', '')}
        t['userType'] = u.get('role', '')
    return t


@router.get('/tickets')
async def admin_list_tickets(status: Optional[str] = None, category: Optional[str] = None,
                             search: Optional[str] = None):
    q: Dict = {}
    if status:
        q['status'] = status
    if category:
        q['category'] = category
    if search:
        rx = {'$regex': search, '$options': 'i'}
        q['$or'] = [{'subject': rx}, {'description': rx}, {'ticket_id': rx}]
    rows = await db.support_tickets.find(q, {'_id': 0}).sort('created_at', -1).to_list(500)
    return [await _enrich_ticket(t) for t in rows]


@router.get('/tickets/{ticket_id}')
async def admin_get_ticket(ticket_id: str):
    t = await db.support_tickets.find_one(
        {'$or': [{'id': ticket_id}, {'ticket_id': ticket_id}]}, {'_id': 0})
    if not t:
        raise HTTPException(404, 'Ticket not found')
    return await _enrich_ticket(t)


@router.post('/tickets/{ticket_id}/notes')
async def admin_add_ticket_note(ticket_id: str, body: TicketNoteIn, admin=Depends(current_user)):
    t = await _find_or_404(db.support_tickets, ticket_id, 'Ticket')
    note = {'by': admin['full_name'], 'by_id': admin['id'], 'at': now(), 'text': body.text}
    await db.support_tickets.update_one({'id': ticket_id}, {'$push': {'notes': note}})
    # The ticket owner sees the response in their notifications.
    if t.get('user_id'):
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()), 'user_id': t['user_id'],
            'title': f"Support update · {t.get('ticket_id', ticket_id)}",
            'body': body.text, 'kind': 'support', 'read': False, 'created_at': now()})
    await write_audit(admin, 'admin.ticket_note',
                      f"Added note to {t.get('ticket_id', ticket_id)}", target_id=ticket_id)
    return {'ok': True, 'id': ticket_id, 'note': {**note, 'at': iso(note['at'])}}


@router.patch('/tickets/{ticket_id}')
async def admin_patch_ticket(ticket_id: str, body: TicketPatch, admin=Depends(current_user)):
    t = await _find_or_404(db.support_tickets, ticket_id, 'Ticket')
    updates = body.model_dump(exclude_none=True)
    if not updates:
        return await _enrich_ticket(t)
    updates['updated_at'] = now()
    await db.support_tickets.update_one({'id': ticket_id}, {'$set': updates})
    if 'status' in updates and t.get('user_id'):
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()), 'user_id': t['user_id'],
            'title': f"Ticket {updates['status'].lower()} · {t.get('ticket_id', ticket_id)}",
            'body': f"Your support ticket is now {updates['status']}.",
            'kind': 'support', 'read': False, 'created_at': now()})
    await write_audit(admin, 'admin.ticket_update',
                      f"Updated {t.get('ticket_id', ticket_id)} "
                      f"({', '.join(f'{k}={v}' for k, v in updates.items() if k != 'updated_at')})",
                      target_id=ticket_id)
    fresh = await db.support_tickets.find_one({'id': ticket_id}, {'_id': 0})
    return await _enrich_ticket(fresh)


# ═════════════════════════════════════════════════════════════════════════════
# SYSTEM ALERTS
# ═════════════════════════════════════════════════════════════════════════════
class AlertResolveIn(BaseModel):
    note: Optional[str] = ''


@router.get('/alerts')
async def admin_list_alerts(status: Optional[str] = None, severity: Optional[str] = None):
    q: Dict = {}
    if status:
        q['status'] = status
    if severity:
        q['severity'] = severity
    return await db.system_alerts.find(q, {'_id': 0}).sort('timestamp', -1).to_list(500)


@router.post('/alerts/{alert_id}/retry')
async def admin_retry_alert(alert_id: str, admin=Depends(current_user)):
    alert = await _find_or_404(db.system_alerts, alert_id, 'Alert')
    await db.system_alerts.update_one({'id': alert_id}, {
        '$set': {'last_retry_at': now()}, '$inc': {'retries': 1}})
    await write_audit(admin, 'admin.alert_retry',
                      f"Retried failed operation for alert: {alert.get('type')}",
                      target_id=alert_id)
    return {'ok': True, 'id': alert_id}


@router.post('/alerts/{alert_id}/notify-user')
async def admin_alert_notify_user(alert_id: str, admin=Depends(current_user)):
    alert = await _find_or_404(db.system_alerts, alert_id, 'Alert')
    affected = (alert.get('affected') or '').strip().lower()
    target = await db.users.find_one({'email': affected}, {'_id': 0, 'id': 1})
    if not target:
        raise HTTPException(404, 'Affected user not found')
    await db.notifications.insert_one({
        'id': str(uuid.uuid4()), 'user_id': target['id'],
        'title': 'Action needed on your account',
        'body': alert.get('description', 'Our team flagged an issue affecting your account.'),
        'kind': 'system', 'read': False, 'created_at': now()})
    await db.system_alerts.update_one({'id': alert_id}, {'$set': {'user_notified_at': now()}})
    await write_audit(admin, 'admin.alert_notify_user',
                      f"Notified {affected} about alert: {alert.get('type')}",
                      target_id=alert_id)
    return {'ok': True, 'id': alert_id, 'notified': affected}


@router.post('/alerts/{alert_id}/escalate')
async def admin_escalate_alert(alert_id: str, admin=Depends(current_user)):
    alert = await _find_or_404(db.system_alerts, alert_id, 'Alert')
    await db.system_alerts.update_one({'id': alert_id}, {'$set': {
        'severity': 'critical', 'escalated': True, 'escalated_at': now(),
        'escalated_by': admin['full_name']}})
    await write_audit(admin, 'admin.alert_escalate',
                      f"Escalated alert: {alert.get('type')}", target_id=alert_id)
    return {'ok': True, 'id': alert_id, 'severity': 'critical'}


@router.post('/alerts/{alert_id}/resolve')
async def admin_resolve_alert(alert_id: str, body: AlertResolveIn, admin=Depends(current_user)):
    alert = await _find_or_404(db.system_alerts, alert_id, 'Alert')
    if alert.get('status') == 'resolved':
        raise HTTPException(400, 'Alert is already resolved')
    await db.system_alerts.update_one({'id': alert_id}, {'$set': {
        'status': 'resolved', 'resolved_at': now(), 'resolved_by': admin['full_name'],
        'resolution_note': body.note or ''}})
    await write_audit(admin, 'admin.alert_resolve',
                      f"Resolved alert: {alert.get('type')}"
                      + (f" — {body.note}" if body.note else ''),
                      target_id=alert_id)
    return {'ok': True, 'id': alert_id, 'status': 'resolved'}


# ═════════════════════════════════════════════════════════════════════════════
# NOTIFICATION MONITORING (delivery log + stats + reminder settings)
# ═════════════════════════════════════════════════════════════════════════════
NOTIF_SETTINGS_KEY = 'notification_settings'
DEFAULT_NOTIF_SETTINGS = {
    'visitReminderHours': 24,
    'medicationReminderMins': 30,
    'channels': {'push': True, 'sms': True, 'email': True},
}


class NotifSettingsPatch(BaseModel):
    visitReminderHours: Optional[int] = Field(None, ge=1, le=168)
    medicationReminderMins: Optional[int] = Field(None, ge=1, le=1440)
    channels: Optional[Dict[str, bool]] = None


def _mask_recipient(rec: str) -> str:
    rec = (rec or '').strip()
    return _mask_email(rec) if '@' in rec else _mask_phone(rec)


@router.get('/notifications/stats')
async def admin_notification_stats():
    total = await db.notification_deliveries.count_documents({})
    by_status: Dict[str, int] = {}
    for st in ('Delivered', 'Failed', 'Pending'):
        by_status[st.lower()] = await db.notification_deliveries.count_documents({'status': st})
    by_channel: Dict[str, int] = {}
    for ch in ('Push', 'SMS', 'Email'):
        by_channel[ch.lower()] = await db.notification_deliveries.count_documents({'channel': ch})
    failures_24h = await db.notification_deliveries.count_documents({
        'status': 'Failed', 'sentAt': {'$gte': now() - timedelta(hours=24)}})
    return {'total': total, 'by_status': by_status, 'by_channel': by_channel,
            'failures_24h': failures_24h}


@router.get('/notifications/log')
async def admin_notification_log(status: Optional[str] = None, channel: Optional[str] = None,
                                 limit: int = Query(200, le=1000)):
    q: Dict = {}
    if status:
        q['status'] = status
    if channel:
        q['channel'] = channel
    rows = await db.notification_deliveries.find(q, {'_id': 0}).sort('sentAt', -1).to_list(limit)
    for r in rows:
        r['recipient'] = _mask_recipient(r.get('recipient', ''))
    return rows


@router.get('/notifications/settings')
async def admin_get_notification_settings():
    doc = await db.app_content.find_one({'key': NOTIF_SETTINGS_KEY}, {'_id': 0, 'key': 0})
    return doc or dict(DEFAULT_NOTIF_SETTINGS)


@router.patch('/notifications/settings')
async def admin_patch_notification_settings(body: NotifSettingsPatch, admin=Depends(current_user)):
    current = await db.app_content.find_one({'key': NOTIF_SETTINGS_KEY}, {'_id': 0, 'key': 0}) \
        or dict(DEFAULT_NOTIF_SETTINGS)
    updates = body.model_dump(exclude_none=True)
    if 'channels' in updates:
        merged = {**current.get('channels', {}), **updates['channels']}
        unknown = set(merged) - set(DEFAULT_NOTIF_SETTINGS['channels'])
        if unknown:
            raise HTTPException(400, f"Unknown channels: {', '.join(sorted(unknown))}")
        updates['channels'] = merged
    merged_doc = {**current, **updates}
    await db.app_content.update_one({'key': NOTIF_SETTINGS_KEY},
                                    {'$set': merged_doc}, upsert=True)
    await write_audit(admin, 'admin.notification_settings',
                      f"Updated notification settings ({', '.join(updates)})",
                      changes=updates)
    return merged_doc


@router.post('/notifications/{delivery_id}/retry')
async def admin_retry_notification(delivery_id: str, admin=Depends(current_user)):
    d = await _find_or_404(db.notification_deliveries, delivery_id, 'Delivery record')
    if d.get('status') != 'Failed':
        raise HTTPException(400, 'Only failed deliveries can be retried')
    await db.notification_deliveries.update_one({'id': delivery_id}, {
        '$set': {'status': 'Pending', 'retried_at': now(), 'error': ''},
        '$inc': {'retries': 1}})
    await write_audit(admin, 'admin.notification_retry',
                      f"Retried {d.get('channel', '')} delivery to "
                      f"{_mask_recipient(d.get('recipient', ''))}",
                      target_id=delivery_id)
    return {'ok': True, 'id': delivery_id, 'status': 'Pending'}


# ═════════════════════════════════════════════════════════════════════════════
# AUDIT LOG (admin scope: unrestricted view + summary + security alerts + export)
# ═════════════════════════════════════════════════════════════════════════════
def _audit_query(category: Optional[str], from_: Optional[str], to: Optional[str],
                 user_id: Optional[str], org: Optional[str], status: Optional[str]) -> Dict:
    f = _parse_ymd(from_, 'from')
    t = _parse_ymd(to, 'to')
    if f is not None and t is not None and t < f:
        raise HTTPException(400, 'to must be on or after from')
    q: Dict = {}
    if category:
        q['category'] = category
    if user_id:
        q['user_id'] = user_id
    if org:
        q['org'] = org
    if status:
        q['status'] = status
    if f is not None or t is not None:
        rng: Dict = {}
        if f is not None:
            rng['$gte'] = datetime(f.year, f.month, f.day, tzinfo=timezone.utc)
        if t is not None:
            rng['$lt'] = datetime(t.year, t.month, t.day, tzinfo=timezone.utc) + timedelta(days=1)
        q['created_at'] = rng
    return q


@router.get('/audit-logs')
async def admin_audit_logs(category: Optional[str] = None,
                           from_: Optional[str] = Query(None, alias='from'),
                           to: Optional[str] = None, user_id: Optional[str] = None,
                           org: Optional[str] = None, status: Optional[str] = None,
                           limit: int = Query(300, le=2000)):
    q = _audit_query(category, from_, to, user_id, org, status)
    return await db.audit_logs.find(q, {'_id': 0}).sort('created_at', -1).to_list(limit)


@router.get('/audit-logs/summary')
async def admin_audit_summary():
    total = await db.audit_logs.count_documents({})
    last_24h = await db.audit_logs.count_documents(
        {'created_at': {'$gte': now() - timedelta(hours=24)}})
    failures_24h = await db.audit_logs.count_documents(
        {'status': 'failure', 'created_at': {'$gte': now() - timedelta(hours=24)}})
    by_category = {g['_id'] or 'other': g['n'] async for g in db.audit_logs.aggregate(
        [{'$group': {'_id': '$category', 'n': {'$sum': 1}}}])}
    return {'total': total, 'last_24h': last_24h, 'failures_24h': failures_24h,
            'by_category': by_category}


@router.get('/audit-logs/security-alerts')
async def admin_audit_security_alerts(threshold: int = Query(3, ge=2, le=20)):
    """Failed-login patterns from the audit trail: any (user, ip) with >=
    `threshold` failures in the last 24h is a security signal."""
    since = now() - timedelta(hours=24)
    rows = await db.audit_logs.find(
        {'category': 'login', 'status': 'failure', 'created_at': {'$gte': since}},
        {'_id': 0}).to_list(5000)
    buckets: Dict[tuple, dict] = {}
    for r in rows:
        key = (r.get('user_id') or r.get('user_name') or 'unknown', r.get('ip', ''))
        b = buckets.setdefault(key, {'user_id': r.get('user_id'),
                                     'user_name': r.get('user_name', ''),
                                     'ip': r.get('ip', ''), 'count': 0,
                                     'last_at': r.get('created_at')})
        b['count'] += 1
        if r.get('created_at') and (not b['last_at'] or r['created_at'] > b['last_at']):
            b['last_at'] = r['created_at']
    alerts = [b for b in buckets.values() if b['count'] >= threshold]
    for a in alerts:
        a['last_at'] = iso(a['last_at'])
        a['pattern'] = 'repeated_failed_login'
    return sorted(alerts, key=lambda a: -a['count'])


@router.get('/audit-logs/export')
async def admin_audit_export(category: Optional[str] = None,
                             from_: Optional[str] = Query(None, alias='from'),
                             to: Optional[str] = None, user_id: Optional[str] = None,
                             org: Optional[str] = None, status: Optional[str] = None,
                             admin=Depends(current_user)):
    q = _audit_query(category, from_, to, user_id, org, status)
    rows = await db.audit_logs.find(q, {'_id': 0}).sort('created_at', -1).to_list(5000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['time', 'user', 'role', 'org', 'category', 'action', 'detail',
                'status', 'ip', 'device'])
    for r in rows:
        w.writerow([iso(r.get('created_at')), r.get('user_name'), r.get('role'),
                    r.get('org'), r.get('category'), r.get('action'), r.get('detail'),
                    r.get('status'), r.get('ip'), r.get('device')])
    await write_audit(admin, 'admin.audit_export', f'Exported {len(rows)} audit rows to CSV')
    return Response(content=buf.getvalue(), media_type='text/csv',
                    headers={'Content-Disposition': 'attachment; filename="audit-logs.csv"'})


# ═════════════════════════════════════════════════════════════════════════════
# ADMIN TRIALS (read-only aggregates; subjects masked unless an active BTG
# session — every unmasked read is audited with the session id)
# ═════════════════════════════════════════════════════════════════════════════
async def _active_btg_session(user_id: str) -> Optional[dict]:
    """The caller's newest ACTIVE break-the-glass session, enforcing the 2h TTL:
    an expired session is tombstoned on sight and never grants access."""
    s = await db.emergency_sessions.find_one(
        {'user_id': user_id, 'status': 'active'}, {'_id': 0},
        sort=[('started_at', -1)])
    if not s:
        return None
    if s['expires_at'] <= now():
        await db.emergency_sessions.update_one(
            {'id': s['id']}, {'$set': {'status': 'expired', 'ended_at': s['expires_at']}})
        return None
    return s


async def _trial_aggregates(trial: dict) -> dict:
    tid = trial['id']
    enrolled = await db.patients.count_documents({'trial_id': tid})
    completed = await db.visit_instances.count_documents(
        {'trial_id': tid, 'status': 'completed'})
    upcoming = await db.visit_instances.count_documents(
        {'trial_id': tid, 'status': 'upcoming'})
    missed = await db.visit_instances.count_documents(
        {'trial_id': tid, 'status': 'missed'})
    creator = await db.users.find_one(
        {'id': trial.get('created_by')},
        {'_id': 0, 'full_name': 1, 'role': 1, 'organization': 1})
    modifier_id = trial.get('updated_by') or trial.get('created_by')
    modifier = creator if modifier_id == trial.get('created_by') else await db.users.find_one(
        {'id': modifier_id},
        {'_id': 0, 'full_name': 1, 'role': 1, 'organization': 1})
    latest_change = await db.audit_logs.find_one(
        {'trial_id': tid, 'action': {'$in': ['trial.create', 'trial.update']}},
        {'_id': 0, 'action': 1, 'detail': 1, 'changes': 1, 'created_at': 1},
        sort=[('created_at', -1)])
    changed_fields = sorted((latest_change or {}).get('changes', {}).keys())
    if changed_fields:
        change_summary = 'Updated ' + ', '.join(
            field.replace('_', ' ') for field in changed_fields)
    elif (latest_change or {}).get('action') == 'trial.create':
        change_summary = 'Trial created'
    else:
        change_summary = (latest_change or {}).get('detail', '')
    creator_role = (creator or {}).get('role', '')
    owner_type = 'CRO' if creator_role == 'cro' else 'Sponsor'
    owner_name = (
        trial.get('cro_name')
        or trial.get('sponsor_name')
        or (creator or {}).get('organization', '')
    )
    return {
        'id': tid, 'title': trial.get('title'), 'protocol_id': trial.get('protocol_id'),
        'phase': trial.get('phase'), 'condition': trial.get('condition'),
        'sponsor': owner_name, 'sponsorOrCroName': owner_name,
        'ownerType': owner_type, 'cro': owner_name if owner_type == 'CRO' else '',
        'status': trial.get('status', 'active'),
        'patients': enrolled, 'targetEnrollment': trial.get('target_enrollment'),
        'scheduleVersion': trial.get('schedule_version', 1),
        'schedule_status': trial.get('schedule_status', ''),
        'visits': {'completed': completed, 'upcoming': upcoming, 'missed': missed},
        'lastModified': iso(trial.get('updated_at') or trial.get('created_at')),
        'modifiedBy': (
            trial.get('updated_by_name')
            or (modifier or {}).get('full_name', '')
            or (creator or {}).get('full_name', '')
        ),
        'modifiedByRole': (modifier or {}).get('role', ''),
        'changeSummary': change_summary,
        'changedFields': changed_fields,
    }


@router.get('/trials')
async def admin_list_trials():
    """Read-only trial monitoring: metadata + enrollment aggregates ONLY —
    no subject rows, no patient PII."""
    trials = await db.trials.find({}, {'_id': 0}).sort('created_at', -1).to_list(500)
    return [await _trial_aggregates(t) for t in trials]


@router.get('/trials/{trial_id}')
async def admin_get_trial(trial_id: str, admin=Depends(current_user)):
    """Trial detail: aggregates + subject list. Subjects are ALWAYS masked
    (SUBJ-xxx + initials) unless the caller holds an active break-the-glass
    session, in which case identified data is returned AND the read itself is
    written to the audit trail with the session id."""
    trial = await _find_or_404(db.trials, trial_id, 'Trial')
    out = await _trial_aggregates(trial)
    patients = await db.patients.find({'trial_id': trial_id}, {'_id': 0}).to_list(1000)
    session = await _active_btg_session(admin['id'])
    if session:
        out['subjects'] = [{
            'subject': f"SUBJ-{(p.get('id') or '')[-3:]}",
            'full_name': p.get('full_name', ''), 'email': p.get('email', ''),
            'status': p.get('status', ''), 'enrolled_date': p.get('enrolled_date', ''),
        } for p in patients]
        out['unmasked'] = True
        out['btg_session_id'] = session['id']
        await write_audit(admin, 'emergency.read',
                          f"Break-the-glass read of identified subjects for trial "
                          f"{trial.get('protocol_id', trial_id)}",
                          category='emergency', target_id=trial_id,
                          btg_session_id=session['id'], trial_id=trial_id)
    else:
        out['subjects'] = [_masked_subject(p) for p in patients]
        out['unmasked'] = False
    return out


# ═════════════════════════════════════════════════════════════════════════════
# TERMS & PRIVACY VERSIONS
# ═════════════════════════════════════════════════════════════════════════════
TERMS_TYPE_TO_LEGAL_KEY = {'ToS': 'terms', 'Privacy': 'privacy'}


def _version_tuple(v: str) -> tuple:
    try:
        return tuple(int(x) for x in str(v).strip().split('.'))
    except (ValueError, AttributeError):
        raise HTTPException(400, 'Version must be numeric, e.g. "2.0" or "2.1.3"')


class TermsVersionIn(BaseModel):
    type: Literal['ToS', 'Privacy']
    version: str = Field(min_length=1)
    content: str = Field(min_length=1)
    effectiveDate: Optional[str] = None
    changeSummary: Optional[str] = ''
    forceReacceptance: bool = False


class TermsVersionPatch(BaseModel):
    content: Optional[str] = None
    changeSummary: Optional[str] = None


@router.get('/terms/versions')
async def admin_terms_versions(type: Optional[Literal['ToS', 'Privacy']] = None):
    q: Dict = {}
    if type:
        q['type'] = type
    rows = await db.terms_versions.find(q, {'_id': 0}).to_list(500)

    def _safe_tuple(r):
        try:
            return _version_tuple(r.get('version', '0'))
        except HTTPException:
            return (0,)
    rows.sort(key=lambda r: (r.get('type', ''), _safe_tuple(r)), reverse=True)
    return rows


@router.post('/terms/versions')
async def admin_publish_terms(body: TermsVersionIn, admin=Depends(current_user)):
    """Publish a new ToS/Privacy version. The version must be strictly greater
    than every existing version of that type; the previous active version is
    superseded; forceReacceptance clears every user's acceptance so the app
    re-prompts on next login. The user-facing /api/legal document is synced."""
    new_v = _version_tuple(body.version)
    existing = await db.terms_versions.find({'type': body.type}, {'_id': 0}).to_list(500)
    for e in existing:
        try:
            old_v = _version_tuple(e.get('version', '0'))
        except HTTPException:
            continue
        if old_v >= new_v:
            raise HTTPException(
                400, f"Version must be greater than the existing {e.get('version')}")
    n = now()
    await db.terms_versions.update_many(
        {'type': body.type, 'status': 'active'},
        {'$set': {'status': 'superseded', 'supersededAt': n}})
    doc = {
        'id': str(uuid.uuid4()), 'type': body.type, 'version': body.version.strip(),
        'status': 'active', 'content': body.content,
        'effectiveDate': body.effectiveDate or n.date().isoformat(),
        'changeSummary': body.changeSummary or '', 'forceReacceptance': body.forceReacceptance,
        'createdAt': n, 'activatedAt': n, 'acceptedBy': 0, 'created_by': admin['id'],
    }
    await db.terms_versions.insert_one(doc)
    # Sync the user-facing legal document (GET /api/legal/{terms|privacy}).
    legal_key = TERMS_TYPE_TO_LEGAL_KEY[body.type]
    await db.app_content.update_one({'key': legal_key}, {'$set': {
        'version': doc['version'], 'effective_date': doc['effectiveDate'],
        'blocks': [{'heading': f"{body.type} v{doc['version']}", 'body': body.content}],
    }}, upsert=True)
    reacceptance_required = 0
    if body.forceReacceptance:
        res = await db.users.update_many(
            {'terms_accepted_at': {'$exists': True}},
            {'$unset': {'terms_accepted_at': ''}})
        reacceptance_required = res.modified_count
    await write_audit(admin, 'admin.terms_publish',
                      f"Published {body.type} v{doc['version']}"
                      + (' (forced re-acceptance)' if body.forceReacceptance else ''),
                      target_id=doc['id'], forceReacceptance=body.forceReacceptance,
                      reacceptance_required=reacceptance_required)
    return {**serialize(doc), 'reacceptance_required': reacceptance_required}


@router.patch('/terms/versions/{version_id}')
async def admin_patch_terms(version_id: str, body: TermsVersionPatch,
                            admin=Depends(current_user)):
    v = await _find_or_404(db.terms_versions, version_id, 'Terms version')
    updates = body.model_dump(exclude_none=True)
    if not updates:
        return v
    updates['updated_at'] = now()
    await db.terms_versions.update_one({'id': version_id}, {'$set': updates})
    if 'content' in updates and v.get('status') == 'active':
        legal_key = TERMS_TYPE_TO_LEGAL_KEY.get(v.get('type'))
        if legal_key:
            await db.app_content.update_one({'key': legal_key}, {'$set': {
                'blocks': [{'heading': f"{v.get('type')} v{v.get('version')}",
                            'body': updates['content']}]}}, upsert=True)
    await write_audit(admin, 'admin.terms_update',
                      f"Edited {v.get('type')} v{v.get('version')} "
                      f"({', '.join(k for k in updates if k != 'updated_at')})",
                      target_id=version_id)
    return await db.terms_versions.find_one({'id': version_id}, {'_id': 0})


@router.get('/terms/acceptances')
async def admin_terms_acceptances():
    rows = await db.users.find({'terms_accepted_at': {'$exists': True}},
                               USER_PROJECTION).sort('terms_accepted_at', -1).to_list(2000)
    out = []
    for u in rows:
        u = _pseudonymize_patient(u)
        out.append({'user_id': u['id'], 'name': u.get('full_name', ''),
                    'email': u.get('email', ''), 'role': u.get('role', ''),
                    'accepted_at': iso(u.get('terms_accepted_at'))})
    return out


# ═════════════════════════════════════════════════════════════════════════════
# REPORTS (approved PDF/Excel formats, stored via storage.py, admin download)
# ═════════════════════════════════════════════════════════════════════════════
class ReportGenerateIn(BaseModel):
    type: Literal['users', 'org-users', 'user-status', 'login-activity', 'trial-summary']
    format: Literal['pdf', 'xlsx'] = 'pdf'
    from_: Optional[str] = Field(None, alias='from')
    to: Optional[str] = None

    model_config = {'populate_by_name': True}


async def _report_rows(rtype: str, from_: Optional[str], to: Optional[str]):
    """Build (headers, rows) for a report type. Patient PII is pseudonymized."""
    if rtype == 'users':
        users = await db.users.find({}, USER_PROJECTION).sort('created_at', -1).to_list(5000)
        rows = []
        for u in users:
            u = _pseudonymize_patient(u)
            rows.append([u.get('id'), u.get('full_name'), u.get('email'), u.get('role'),
                         u.get('organization'), _user_status(u), iso(u.get('created_at'))])
        return ['id', 'name', 'email', 'role', 'organization', 'status', 'created_at'], rows
    if rtype == 'org-users':
        users = await db.users.find({}, {'_id': 0, 'organization': 1, 'role': 1}).to_list(10000)
        roles = ['sponsor', 'cro', 'smo', 'site', 'pi', 'crc', 'patient', 'admin']
        per_org: Dict[str, Dict[str, int]] = {}
        for u in users:
            org = u.get('organization') or '(none)'
            b = per_org.setdefault(org, {r: 0 for r in roles})
            if u.get('role') in b:
                b[u['role']] += 1
        rows = [[org, sum(counts.values())] + [counts[r] for r in roles]
                for org, counts in sorted(per_org.items())]
        return ['organization', 'total'] + roles, rows
    if rtype == 'user-status':
        users = await db.users.find({}, {'_id': 0, 'status': 1, 'lock_info': 1}).to_list(10000)
        counts: Dict[str, int] = {}
        for u in users:
            counts[_user_status(u)] = counts.get(_user_status(u), 0) + 1
        return ['status', 'count'], [[k, v] for k, v in sorted(counts.items())]
    if rtype == 'login-activity':
        q = _audit_query('login', from_, to, None, None, None)
        logs = await db.audit_logs.find(q, {'_id': 0}).sort('created_at', -1).to_list(5000)
        rows = [[iso(r.get('created_at')), r.get('user_name'), r.get('role'),
                 r.get('action'), r.get('status'), r.get('ip')] for r in logs]
        return ['time', 'user', 'role', 'action', 'status', 'ip'], rows
    # trial-summary
    trials = await db.trials.find({}, {'_id': 0}).to_list(500)
    rows = []
    for t in trials:
        agg = await _trial_aggregates(t)
        rows.append([agg['protocol_id'], agg['title'], agg['phase'], agg['status'],
                     agg['patients'], agg['visits']['completed'], agg['visits']['missed']])
    return ['protocol', 'title', 'phase', 'status', 'enrolled',
            'visits_completed', 'visits_missed'], rows


def _report_bytes(
        title: str, headers: List[str], rows: List[List], fmt: str) -> tuple[bytes, str, str]:
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Report'
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='A6213F')
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            sheet.append(['' if value is None else value for value in row])
        sheet.freeze_panes = 'A2'
        for column in sheet.columns:
            values = [str(cell.value or '') for cell in column]
            width = min(max(max((len(value) for value in values), default=8) + 2, 10), 42)
            sheet.column_dimensions[column[0].column_letter].width = width
        workbook.save(output)
        return (
            output.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xlsx',
        )

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        rightMargin=10 * mm, leftMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    table_data = [headers] + [[str(value if value is not None else '') for value in row]
                              for row in rows]
    column_width = (landscape(A4)[0] - 20 * mm) / max(len(headers), 1)
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[column_width] * len(headers),
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#A6213F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E6D6C5')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white, colors.HexColor('#FEFAF1')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    document.build([
        Paragraph(title, styles['Title']),
        Spacer(1, 5 * mm),
        table,
    ])
    return output.getvalue(), 'application/pdf', 'pdf'


@router.post('/reports/generate')
async def admin_generate_report(body: ReportGenerateIn, admin=Depends(current_user)):
    headers, rows = await _report_rows(body.type, body.from_, body.to)
    n = now()
    title = f"{body.type.replace('-', ' ').title()} report"
    data, content_type, extension = _report_bytes(
        title, headers, rows, body.format)
    key = f'reports/{uuid.uuid4()}.{extension}'
    await file_storage.get_storage().save(key, data, content_type)
    doc = {
        'id': str(uuid.uuid4()), 'type': body.type,
        'name': f"{body.type}-{n.strftime('%Y%m%d-%H%M%S')}.{extension}",
        'format': extension, 'content_type': content_type,
        'key': key, 'size': len(data), 'rows': len(rows),
        'params': {'from': body.from_, 'to': body.to},
        'created_by': admin['id'], 'created_by_name': admin['full_name'],
        'created_at': n,
    }
    await db.admin_reports.insert_one(doc)
    await write_audit(admin, 'admin.report_generate',
                      f"Generated {body.type} {extension.upper()} report ({len(rows)} rows)",
                      target_id=doc['id'])
    return {**serialize(doc), 'download_url': f"/api/admin/reports/{doc['id']}/download"}


@router.get('/reports/recent')
async def admin_recent_reports():
    rows = await db.admin_reports.find({}, {'_id': 0}).sort('created_at', -1).to_list(20)
    for r in rows:
        r['download_url'] = f"/api/admin/reports/{r['id']}/download"
    return rows


@router.get('/reports/{report_id}/download')
async def admin_download_report(report_id: str, admin=Depends(current_user)):
    rep = await _find_or_404(db.admin_reports, report_id, 'Report')
    try:
        data, _ct = await file_storage.get_storage().open(rep['key'])
    except FileNotFoundError:
        raise HTTPException(404, 'Report file is missing')
    await write_audit(admin, 'admin.report_download',
                      f"Downloaded report {rep.get('name')}", target_id=report_id)
    return Response(content=data, media_type=rep.get('content_type') or (
                        'application/pdf' if rep.get('format') == 'pdf'
                        else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    headers={'Content-Disposition':
                             f"attachment; filename=\"{rep.get('name', 'report')}\""})


# ═════════════════════════════════════════════════════════════════════════════
# ORGANIZATION TRIAL-CREATION DELEGATION REQUESTS
# ═════════════════════════════════════════════════════════════════════════════
class OrgDelegationDecisionIn(BaseModel):
    reason: Optional[str] = Field('', max_length=1000)


@router.get('/org-delegation-requests')
async def admin_list_org_delegation_requests(status: Optional[str] = None):
    q: Dict = {}
    if status:
        if status not in ('pending', 'approved', 'rejected'):
            raise HTTPException(400, 'Invalid delegation-request status')
        q['status'] = status
    return await db.org_delegation_requests.find(
        q, {'_id': 0}).sort('created_at', -1).to_list(500)


async def _decide_org_delegation_request(request_id: str, decision: str,
                                         body: OrgDelegationDecisionIn,
                                         admin: dict) -> dict:
    req = await _find_or_404(
        db.org_delegation_requests, request_id, 'Delegation request')
    if req.get('status') != 'pending':
        raise HTTPException(400, f"This request is already {req.get('status')}")
    org = await db.organizations.find_one({'id': req.get('org_id')}, {'_id': 0})
    if not org:
        raise HTTPException(404, 'Organization not found')
    if org.get('type') not in ('site', 'smo'):
        raise HTTPException(
            400, 'Trial-creation delegation is only available to Site and SMO organizations')

    n = now()
    updates = {
        'status': decision,
        'decided_by': admin['id'],
        'decider_name': admin.get('full_name', ''),
        'decision_reason': (body.reason or '').strip(),
        'decided_at': n,
    }
    result = await db.org_delegation_requests.update_one(
        {'id': request_id, 'status': 'pending'}, {'$set': updates})
    if not result.modified_count:
        raise HTTPException(409, 'This request was already actioned')

    delegated = decision == 'approved'
    await db.organizations.update_one(
        {'id': org['id']},
        {'$set': {
            'trial_creation_delegated': delegated,
            'trial_creation_delegation_request_id': request_id,
            'trial_creation_delegation_updated_at': n,
            'trial_creation_delegation_updated_by': admin['id'],
        }})
    if req.get('requested_by'):
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()), 'user_id': req['requested_by'],
            'title': f"Trial creation request {decision}",
            'body': (
                f"Trial-creation delegation for {org.get('name', 'your organization')} "
                f"was {decision}."
                + (f" {(body.reason or '').strip()}" if (body.reason or '').strip() else '')
            ),
            'kind': 'system', 'read': False, 'created_at': n,
        })
    await write_audit(
        admin,
        f'admin.org_delegation_{decision}',
        f"{decision.title()} trial-creation delegation for {org.get('name')}"
        + (f" — {(body.reason or '').strip()}" if (body.reason or '').strip() else ''),
        target_id=request_id,
        org_id=org['id'],
        changes={'trial_creation_delegated': delegated},
    )
    fresh = await db.org_delegation_requests.find_one(
        {'id': request_id}, {'_id': 0})
    return serialize(fresh)


@router.post('/org-delegation-requests/{request_id}/approve')
async def admin_approve_org_delegation_request(
        request_id: str, body: OrgDelegationDecisionIn,
        admin=Depends(current_user)):
    return await _decide_org_delegation_request(
        request_id, 'approved', body, admin)


@router.post('/org-delegation-requests/{request_id}/reject')
async def admin_reject_org_delegation_request(
        request_id: str, body: OrgDelegationDecisionIn,
        admin=Depends(current_user)):
    return await _decide_org_delegation_request(
        request_id, 'rejected', body, admin)


# ═════════════════════════════════════════════════════════════════════════════
# ADMIN-STAFF DELEGATIONS
# ═════════════════════════════════════════════════════════════════════════════
DELEGATION_TASKS = ('user_management', 'support_tickets', 'invitations', 'master_data',
                    'notifications', 'reports', 'audit_review')
DelegationTask = Literal['user_management', 'support_tickets', 'invitations', 'master_data',
                         'notifications', 'reports', 'audit_review']


class DelegationIn(BaseModel):
    user_id: str
    tasks: List[DelegationTask] = Field(min_length=1)
    reason: str = Field(min_length=20)


class DelegationPatch(BaseModel):
    tasks: Optional[List[DelegationTask]] = Field(None, min_length=1)
    reason: Optional[str] = Field(None, min_length=20)


@router.get('/delegations')
async def admin_list_delegations(status: Optional[str] = None):
    q: Dict = {}
    if status:
        q['status'] = status
    return await db.admin_delegations.find(q, {'_id': 0}).sort('delegatedDate', -1).to_list(200)


@router.post('/delegations')
async def admin_create_delegation(body: DelegationIn, admin=Depends(current_user)):
    target = await db.users.find_one({'id': body.user_id}, USER_PROJECTION)
    if not target:
        raise HTTPException(404, 'User not found')
    if target['role'] == 'patient':
        raise HTTPException(400, 'Admin tasks cannot be delegated to a patient account')
    existing = await db.admin_delegations.find_one(
        {'user_id': body.user_id, 'status': {'$in': ['active', 'suspended']}})
    if existing:
        raise HTTPException(400, 'This user already has a delegation — edit it instead')
    doc = {
        'id': str(uuid.uuid4()), 'user_id': target['id'],
        'name': target.get('full_name', ''), 'email': target.get('email', ''),
        'tasks': sorted(set(body.tasks)), 'status': 'active', 'reason': body.reason,
        'delegated_by': admin['id'], 'delegatedDate': now(), 'lastActive': None,
    }
    await db.admin_delegations.insert_one(doc)
    await write_audit(admin, 'admin.delegation_create',
                      f"Delegated {', '.join(doc['tasks'])} to {doc['email']} — {body.reason}",
                      target_id=doc['id'])
    return serialize({**doc})


@router.patch('/delegations/{delegation_id}')
async def admin_patch_delegation(delegation_id: str, body: DelegationPatch,
                                 admin=Depends(current_user)):
    d = await _find_or_404(db.admin_delegations, delegation_id, 'Delegation')
    if d.get('status') == 'revoked':
        raise HTTPException(400, 'A revoked delegation cannot be edited')
    updates = body.model_dump(exclude_none=True)
    if not updates:
        return d
    if 'tasks' in updates:
        updates['tasks'] = sorted(set(updates['tasks']))
    updates['updated_at'] = now()
    await db.admin_delegations.update_one({'id': delegation_id}, {'$set': updates})
    await write_audit(admin, 'admin.delegation_update',
                      f"Updated delegation for {d.get('email')}",
                      target_id=delegation_id,
                      changes={k: v for k, v in updates.items() if k != 'updated_at'})
    return await db.admin_delegations.find_one({'id': delegation_id}, {'_id': 0})


@router.post('/delegations/{delegation_id}/suspend')
async def admin_suspend_delegation(delegation_id: str, admin=Depends(current_user)):
    d = await _find_or_404(db.admin_delegations, delegation_id, 'Delegation')
    if d.get('status') != 'active':
        raise HTTPException(400, 'Only an active delegation can be suspended')
    await db.admin_delegations.update_one(
        {'id': delegation_id}, {'$set': {'status': 'suspended', 'suspended_at': now()}})
    await write_audit(admin, 'admin.delegation_suspend',
                      f"Suspended delegation for {d.get('email')}", target_id=delegation_id)
    return {'ok': True, 'id': delegation_id, 'status': 'suspended'}


@router.delete('/delegations/{delegation_id}')
async def admin_revoke_delegation(delegation_id: str, admin=Depends(current_user)):
    d = await _find_or_404(db.admin_delegations, delegation_id, 'Delegation')
    if d.get('status') == 'revoked':
        raise HTTPException(400, 'This delegation is already revoked')
    # Revocation keeps the record (regulated audit trail) — never a hard delete.
    await db.admin_delegations.update_one(
        {'id': delegation_id}, {'$set': {'status': 'revoked', 'revoked_at': now()}})
    await write_audit(admin, 'admin.delegation_revoke',
                      f"Revoked delegation for {d.get('email')}", target_id=delegation_id)
    return {'ok': True, 'id': delegation_id, 'status': 'revoked'}


# ═════════════════════════════════════════════════════════════════════════════
# EMERGENCY ACCESS (Break-The-Glass): request → senior approval → 2h session
# ═════════════════════════════════════════════════════════════════════════════
BTG_SESSION_HOURS = 2


class EmergencyRequestIn(BaseModel):
    reason_category: Literal['patient_safety', 'regulatory_audit', 'data_correction',
                             'incident_investigation', 'other']
    reason_text: str = Field(min_length=10)
    trial_id: Optional[str] = None


class EmergencyDenyIn(BaseModel):
    reason: Optional[str] = ''


@router.post('/emergency/requests')
async def btg_create_request(body: EmergencyRequestIn, admin=Depends(current_user)):
    doc = {
        'id': str(uuid.uuid4()), 'requested_by': admin['id'],
        'requester_name': admin['full_name'], 'reason_category': body.reason_category,
        'reason_text': body.reason_text, 'trial_id': body.trial_id,
        'status': 'pending', 'created_at': now(), 'session_id': None,
    }
    await db.emergency_requests.insert_one(doc)
    await write_audit(admin, 'emergency.request',
                      f"Requested break-the-glass access ({body.reason_category}) — "
                      f"{body.reason_text}", category='emergency', target_id=doc['id'])
    return serialize({**doc})


@router.get('/emergency/requests')
async def btg_list_requests(
        status: Optional[str] = 'pending', admin=Depends(current_user)):
    if status and status not in ('pending', 'approved', 'denied'):
        raise HTTPException(400, 'Invalid emergency-request status')
    q: Dict = {}
    if status:
        q['status'] = status
    rows = await db.emergency_requests.find(
        q, {'_id': 0}).sort('created_at', -1).to_list(500)
    return [serialize({
        **row,
        'can_action': (
            row.get('status') == 'pending'
            and row.get('requested_by') != admin['id']
        ),
        'is_own_request': row.get('requested_by') == admin['id'],
    }) for row in rows]


@router.get('/emergency/requests/{request_id}')
async def btg_get_request(request_id: str):
    req = await _find_or_404(db.emergency_requests, request_id, 'Emergency request')
    session = None
    if req.get('session_id'):
        session = await db.emergency_sessions.find_one({'id': req['session_id']}, {'_id': 0})
        if session and session.get('status') == 'active' and session['expires_at'] <= now():
            await db.emergency_sessions.update_one(
                {'id': session['id']},
                {'$set': {'status': 'expired', 'ended_at': session['expires_at']}})
            session['status'] = 'expired'
    return {**req, 'session': session}


@router.post('/emergency/requests/{request_id}/approve')
async def btg_approve_request(request_id: str, admin=Depends(current_user)):
    """Senior approval: a second admin must approve — never the requester
    (two-person rule). Opens a time-boxed 2h session."""
    req = await _find_or_404(db.emergency_requests, request_id, 'Emergency request')
    if req.get('status') != 'pending':
        raise HTTPException(400, f"This request is already {req.get('status')}")
    if req.get('requested_by') == admin['id']:
        raise HTTPException(403, 'You cannot approve your own emergency-access request')
    n = now()
    session = {
        'id': str(uuid.uuid4()), 'request_id': request_id,
        'user_id': req['requested_by'], 'approved_by': admin['id'],
        'approver_name': admin['full_name'], 'status': 'active',
        'started_at': n, 'expires_at': n + timedelta(hours=BTG_SESSION_HOURS),
    }
    await db.emergency_sessions.insert_one(session)
    await db.emergency_requests.update_one({'id': request_id}, {'$set': {
        'status': 'approved', 'approved_by': admin['id'], 'approved_at': n,
        'session_id': session['id']}})
    await write_audit(admin, 'emergency.approve',
                      f"Approved break-the-glass request by {req.get('requester_name')} "
                      f"— session expires {iso(session['expires_at'])}",
                      category='emergency', target_id=request_id,
                      btg_session_id=session['id'])
    return {'ok': True, 'request_id': request_id, 'session': serialize({**session})}


@router.post('/emergency/requests/{request_id}/deny')
async def btg_deny_request(request_id: str, body: EmergencyDenyIn, admin=Depends(current_user)):
    req = await _find_or_404(db.emergency_requests, request_id, 'Emergency request')
    if req.get('status') != 'pending':
        raise HTTPException(400, f"This request is already {req.get('status')}")
    if req.get('requested_by') == admin['id']:
        raise HTTPException(403, 'You cannot action your own emergency-access request')
    await db.emergency_requests.update_one({'id': request_id}, {'$set': {
        'status': 'denied', 'denied_by': admin['id'], 'denied_at': now(),
        'deny_reason': body.reason or ''}})
    await write_audit(admin, 'emergency.deny',
                      f"Denied break-the-glass request by {req.get('requester_name')}"
                      + (f" — {body.reason}" if body.reason else ''),
                      category='emergency', target_id=request_id)
    return {'ok': True, 'request_id': request_id, 'status': 'denied'}


@router.post('/emergency/sessions/{session_id}/end')
async def btg_end_session(session_id: str, admin=Depends(current_user)):
    s = await _find_or_404(db.emergency_sessions, session_id, 'Emergency session')
    if s.get('status') != 'active':
        raise HTTPException(400, f"This session is already {s.get('status')}")
    await db.emergency_sessions.update_one({'id': session_id}, {'$set': {
        'status': 'ended', 'ended_at': now(), 'ended_by': admin['id']}})
    await write_audit(admin, 'emergency.end',
                      'Ended break-the-glass session', category='emergency',
                      target_id=session_id, btg_session_id=session_id)
    return {'ok': True, 'id': session_id, 'status': 'ended'}


@router.get('/emergency/sessions/{session_id}/log')
async def btg_session_log(session_id: str):
    """Every audited action tied to this session (unmasked reads, approve, end)."""
    await _find_or_404(db.emergency_sessions, session_id, 'Emergency session')
    return await db.audit_logs.find({'btg_session_id': session_id}, {'_id': 0}) \
        .sort('created_at', 1).to_list(1000)


# ═════════════════════════════════════════════════════════════════════════════
# MESSAGES (admin broadcasts → fan out to the notifications collection)
# ═════════════════════════════════════════════════════════════════════════════
class BroadcastIn(BaseModel):
    type: Literal['general', 'compliance', 'system', 'targeted', 'urgent'] = 'general'
    subject: str = Field(min_length=1, max_length=120)
    body: str = Field(min_length=1, max_length=2000)
    target: str = 'all'
    allowReplies: bool = True
    scheduleAt: Optional[str] = None


class ReplyRespondIn(BaseModel):
    text: str = Field(min_length=1)


async def _resolve_recipient_ids(target: str) -> List[str]:
    """Resolve a broadcast target expression to user ids WITHOUT sending.
    Forms: all | role:<role> | org:<name> | site:<name> | entity:<orgType> |
    trial:<trial_id> | user:<user_id>."""
    target = (target or 'all').strip()
    if target == 'all':
        rows = await db.users.find({}, {'_id': 0, 'id': 1}).to_list(20000)
        return [r['id'] for r in rows]
    if ':' not in target:
        raise HTTPException(400, 'Invalid target — expected "all" or "<kind>:<value>"')
    kind, value = target.split(':', 1)
    kind, value = kind.strip().lower(), value.strip()
    if not value:
        raise HTTPException(400, 'Target value is required')
    if kind == 'role':
        rows = await db.users.find({'role': value}, {'_id': 0, 'id': 1}).to_list(20000)
        return [r['id'] for r in rows]
    if kind in ('org', 'site'):
        rows = await db.users.find({'organization': value}, {'_id': 0, 'id': 1}).to_list(20000)
        return [r['id'] for r in rows]
    if kind == 'entity':
        if value not in ORG_TYPES:
            raise HTTPException(400, f'Unknown entity type "{value}"')
        orgs = await db.organizations.find({'type': value}, {'_id': 0, 'name': 1}).to_list(2000)
        names = [o['name'] for o in orgs]
        rows = await db.users.find({'organization': {'$in': names}},
                                   {'_id': 0, 'id': 1}).to_list(20000)
        return [r['id'] for r in rows]
    if kind == 'trial':
        trial = await db.trials.find_one({'id': value}, {'_id': 0})
        if not trial:
            raise HTTPException(404, 'Trial not found')
        ids = set()
        if trial.get('created_by'):
            ids.add(trial['created_by'])
        async for p in db.patients.find({'trial_id': value}, {'_id': 0}):
            for k in ('user_id', 'pi_id', 'crc_id'):
                if p.get(k):
                    ids.add(p[k])
        return sorted(ids)
    if kind == 'user':
        u = await db.users.find_one({'id': value}, {'_id': 0, 'id': 1})
        if not u:
            raise HTTPException(404, 'Target user not found')
        return [value]
    raise HTTPException(400, f'Unknown target kind "{kind}"')


@router.get('/messages/recipient-count')
async def broadcast_recipient_count(target: str = 'all'):
    """Recipient preview: resolves the audience WITHOUT sending anything."""
    ids = await _resolve_recipient_ids(target)
    return {'target': target, 'count': len(ids)}


@router.post('/messages')
async def create_broadcast(body: BroadcastIn, admin=Depends(current_user)):
    recipient_ids = await _resolve_recipient_ids(body.target)
    n = now()
    schedule_at = None
    if body.scheduleAt:
        try:
            schedule_at = datetime.fromisoformat(body.scheduleAt.replace('Z', '+00:00'))
            if schedule_at.tzinfo is None:
                schedule_at = schedule_at.replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(400, 'scheduleAt must be an ISO-8601 timestamp')
    doc = {
        'id': str(uuid.uuid4()), 'type': body.type, 'subject': body.subject,
        'body': body.body, 'target': body.target, 'allowReplies': body.allowReplies,
        'scheduleAt': schedule_at, 'created_by': admin['id'],
        'created_by_name': admin['full_name'], 'created_at': n,
        'recipients_count': len(recipient_ids),
    }
    if schedule_at and schedule_at > n:
        doc.update({'status': 'scheduled', 'sent_at': None})
        await db.broadcast_messages.insert_one(doc)
        await write_audit(admin, 'admin.broadcast_schedule',
                          f"Scheduled broadcast \"{body.subject}\" to {len(recipient_ids)} "
                          f"recipients at {iso(schedule_at)}", target_id=doc['id'])
        return serialize({**doc})
    doc.update({'status': 'sent', 'sent_at': n})
    await db.broadcast_messages.insert_one(doc)
    if recipient_ids:
        await db.notifications.insert_many([{
            'id': str(uuid.uuid4()), 'user_id': uid, 'title': body.subject,
            'body': body.body, 'kind': 'broadcast', 'broadcast_id': doc['id'],
            'read': False, 'created_at': n,
        } for uid in recipient_ids])
        await db.notification_deliveries.insert_many([{
            'id': str(uuid.uuid4()), 'type': 'broadcast', 'channel': 'Push',
            'recipient': uid, 'message': body.subject, 'status': 'Delivered',
            'sentAt': n, 'deliveredAt': n, 'error': '', 'broadcast_id': doc['id'],
        } for uid in recipient_ids])
    await write_audit(admin, 'admin.broadcast_send',
                      f"Sent broadcast \"{body.subject}\" to {len(recipient_ids)} recipients "
                      f"(target {body.target})", target_id=doc['id'])
    return serialize({**doc})


@router.get('/messages')
async def list_broadcasts(box: str = 'sent'):
    q: Dict = {}
    if box == 'sent':
        q['status'] = {'$in': ['sent', 'scheduled']}
    rows = await db.broadcast_messages.find(q, {'_id': 0}).sort('created_at', -1).to_list(200)
    for b in rows:
        total = b.get('recipients_count', 0)
        read = await db.notifications.count_documents(
            {'broadcast_id': b['id'], 'read': True}) if total else 0
        replies = await db.broadcast_replies.count_documents({'broadcast_id': b['id']})
        b['read_count'] = read
        b['replies_count'] = replies
    return rows


@router.get('/messages/{message_id}/replies')
async def broadcast_replies(message_id: str):
    await _find_or_404(db.broadcast_messages, message_id, 'Broadcast')
    return await db.broadcast_replies.find({'broadcast_id': message_id}, {'_id': 0}) \
        .sort('created_at', 1).to_list(500)


@router.post('/messages/replies/{reply_id}/respond')
async def broadcast_reply_respond(reply_id: str, body: ReplyRespondIn,
                                  admin=Depends(current_user)):
    reply = await _find_or_404(db.broadcast_replies, reply_id, 'Reply')
    response = {'by': admin['full_name'], 'by_id': admin['id'], 'at': now(),
                'text': body.text}
    await db.broadcast_replies.update_one({'id': reply_id}, {
        '$push': {'responses': response}, '$set': {'status': 'responded'}})
    if reply.get('user_id'):
        await db.notifications.insert_one({
            'id': str(uuid.uuid4()), 'user_id': reply['user_id'],
            'title': 'Reply from the MTB admin team', 'body': body.text,
            'kind': 'broadcast', 'read': False, 'created_at': now()})
    await write_audit(admin, 'admin.broadcast_respond',
                      'Responded to a broadcast reply', target_id=reply_id)
    return {'ok': True, 'id': reply_id, 'status': 'responded'}


@router.post('/messages/replies/{reply_id}/resolve')
async def broadcast_reply_resolve(reply_id: str, admin=Depends(current_user)):
    await _find_or_404(db.broadcast_replies, reply_id, 'Reply')
    await db.broadcast_replies.update_one({'id': reply_id}, {'$set': {
        'status': 'resolved', 'resolved_at': now(), 'resolved_by': admin['id']}})
    await write_audit(admin, 'admin.broadcast_resolve',
                      'Resolved a broadcast reply thread', target_id=reply_id)
    return {'ok': True, 'id': reply_id, 'status': 'resolved'}


# ── Scheduled-broadcast delivery worker ──────────────────────────────────────
# Scheduled broadcasts are stored with status='scheduled' and delivered by this
# worker when scheduleAt comes due. Delivery is exactly-once under concurrency
# and restarts:
#   * a broadcast is CLAIMED atomically (scheduled → sending) so two workers
#     can never both deliver it;
#   * fan-out is idempotent — any partial notification/delivery rows from a
#     crashed attempt are removed for that broadcast before re-inserting;
#   * a 'sending' row whose claim is older than the stale window is re-claimed,
#     so a crash mid-fan-out self-heals instead of stranding the broadcast.
BROADCAST_WORKER_INTERVAL_SEC = int(os.environ.get('BROADCAST_WORKER_INTERVAL_SEC', '30'))
BROADCAST_CLAIM_STALE_SEC = int(os.environ.get('BROADCAST_CLAIM_STALE_SEC', '300'))

_worker_log = logging.getLogger('broadcast_worker')


async def _fan_out_broadcast(doc: dict) -> int:
    """Deliver one claimed broadcast. Safe to re-run for the same broadcast."""
    recipient_ids = await _resolve_recipient_ids(doc.get('target') or 'all')
    sent_at = now()
    # Idempotency: clear any partial rows from an earlier crashed attempt.
    await db.notifications.delete_many({'broadcast_id': doc['id']})
    await db.notification_deliveries.delete_many({'broadcast_id': doc['id']})
    if recipient_ids:
        await db.notifications.insert_many([{
            'id': str(uuid.uuid4()), 'user_id': uid, 'title': doc['subject'],
            'body': doc['body'], 'kind': 'broadcast', 'broadcast_id': doc['id'],
            'read': False, 'created_at': sent_at,
        } for uid in recipient_ids])
        await db.notification_deliveries.insert_many([{
            'id': str(uuid.uuid4()), 'type': 'broadcast', 'channel': 'Push',
            'recipient': uid, 'message': doc['subject'], 'status': 'Delivered',
            'sentAt': sent_at, 'deliveredAt': sent_at, 'error': '',
            'broadcast_id': doc['id'],
        } for uid in recipient_ids])
    await db.broadcast_messages.update_one({'id': doc['id']}, {'$set': {
        'status': 'sent', 'sent_at': sent_at,
        'recipients_count': len(recipient_ids)}})
    actor = {'id': doc.get('created_by'),
             'full_name': doc.get('created_by_name', 'Platform admin'),
             'role': 'admin'}
    await write_audit(actor, 'admin.broadcast_send',
                      f"Delivered scheduled broadcast \"{doc['subject']}\" to "
                      f"{len(recipient_ids)} recipients (target {doc.get('target')})",
                      target_id=doc['id'], scheduled=True)
    return len(recipient_ids)


async def deliver_due_broadcasts(only_id: Optional[str] = None) -> int:
    """Claim and deliver every due scheduled broadcast. Returns count delivered.
    `only_id` narrows the pass to one broadcast (used by tests)."""
    delivered = 0
    while True:
        n = now()
        stale_before = n - timedelta(seconds=BROADCAST_CLAIM_STALE_SEC)
        due: Dict = {'$or': [
            {'status': 'scheduled', 'scheduleAt': {'$lte': n}},
            # self-heal: re-claim a delivery that crashed mid-fan-out
            {'status': 'sending', 'claimed_at': {'$lte': stale_before}},
        ]}
        if only_id:
            due['id'] = only_id
        doc = await db.broadcast_messages.find_one_and_update(
            due,
            {'$set': {'status': 'sending', 'claimed_at': n}},
            projection={'_id': 0},
            return_document=ReturnDocument.AFTER,
        )
        if not doc:
            return delivered
        try:
            count = await _fan_out_broadcast(doc)
            delivered += 1
            _worker_log.info('Delivered scheduled broadcast %s to %d recipients',
                             doc['id'], count)
        except Exception:
            # Keep status='sending' with the current claim time — it will be
            # retried after the stale window instead of hot-looping.
            _worker_log.exception('Scheduled broadcast %s delivery failed; '
                                  'will retry after %ss', doc['id'],
                                  BROADCAST_CLAIM_STALE_SEC)
            return delivered


async def broadcast_worker_loop():
    """Long-running startup task: poll for due scheduled broadcasts."""
    while True:
        try:
            await deliver_due_broadcasts()
        except Exception:
            _worker_log.exception('Broadcast worker pass failed')
        await asyncio.sleep(BROADCAST_WORKER_INTERVAL_SEC)
