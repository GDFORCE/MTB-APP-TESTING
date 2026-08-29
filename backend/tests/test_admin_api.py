"""Platform-admin API — Task 6.1 (SECURITY).

Covers the /api/admin surface: role guard (every group 403s non-admins),
user suspend/unlock/reset/force-logout, org create/patch/merge/name-requests,
master-data approve/edit-approve/reject, admin invitations, and (later groups)
tickets, alerts, notification monitoring, audit views, reports, delegations,
break-the-glass emergency access, broadcasts, and admin trial reads.

Same harness as test_authz_scoping.py: in-process ASGITransport against the
real Atlas DB, RUN_ID-marked data, a single module-level event loop (Motor pins
its io_loop on first use — never asyncio.run here), module teardown cleanup.

Admins cannot self-register (POST /auth/register rejects role=admin), so the
admin actors are inserted directly into the users collection and then logged
in through the real /api/auth/login endpoint.
"""
import asyncio
import io
import sys
import uuid
from datetime import timedelta
from pathlib import Path

import pytest

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

import httpx  # noqa: E402
import server  # noqa: E402

RUN_ID = uuid.uuid4().hex[:8]
PASSWORD = 'Password1!'

LOOP = asyncio.new_event_loop()

_org_ids = []
_extra_cleanup_ids = {'name_requests': [], 'submissions': [], 'trials': [],
                      'reports': [], 'sessions': []}
TERMS_TEST_MAJOR = int(RUN_ID[:6], 16) + 1000   # huge numeric major → always > current


def run(coro):
    return LOOP.run_until_complete(coro)


def make_client():
    return httpx.AsyncClient(
        transport=httpx.ASGITransport(app=server.app), base_url='http://testserver'
    )


async def _login(email):
    async with make_client() as cli:
        r = await cli.post('/api/auth/login', json={'email': email, 'password': PASSWORD})
    assert r.status_code == 200, r.text
    j = r.json()
    return j['user'], {'Authorization': f"Bearer {j['access_token']}"}


async def _make_admin(tag='a'):
    """Admins cannot self-register — insert directly, then real login."""
    email = f'adm-{RUN_ID}-{tag}-{uuid.uuid4().hex[:4]}@example.com'
    doc = {
        'id': str(uuid.uuid4()), 'email': email,
        'full_name': f'Adm {tag.upper()} {RUN_ID}', 'role': 'admin',
        'organization': 'MTB Health Technologies', 'phone': '+91 90000 00000',
        'hashed_password': server.pwd_ctx.hash(PASSWORD),
        'security_question': '', 'security_answer_hash': '',
        'avatar_initials': 'AD', 'created_at': server.now(), 'is_online': False,
    }
    await server.db.users.insert_one(doc)
    return await _login(email)


async def _register(role, org=None):
    email = f'adm-{RUN_ID}-{role}-{uuid.uuid4().hex[:6]}@example.com'
    async with make_client() as cli:
        r = await cli.post('/api/auth/register', json={
            'email': email, 'password': PASSWORD,
            'full_name': f'Test {role.upper()} {RUN_ID}',
            'role': role, 'organization': org,
        })
    assert r.status_code == 200, r.text
    j = r.json()
    return j['user'], {'Authorization': f"Bearer {j['access_token']}"}


@pytest.fixture(scope='module', autouse=True)
def _cleanup():
    yield
    async def clean():
        db = server.db
        test_users = await db.users.find(
            {'email': {'$regex': f'{RUN_ID}'}}, {'_id': 0, 'id': 1}).to_list(1000)
        await db.refresh_tokens.delete_many(
            {'user_id': {'$in': [user['id'] for user in test_users]}})
        await db.users.delete_many({'email': {'$regex': f'{RUN_ID}'}})
        await db.password_reset_tokens.delete_many({'email': {'$regex': RUN_ID}})
        await db.organizations.delete_many({'name': {'$regex': RUN_ID}})
        await db.invitations.delete_many({'email': {'$regex': RUN_ID}})
        await db.org_name_requests.delete_many({'id': {'$in': _extra_cleanup_ids['name_requests']}})
        await db.master_data_submissions.delete_many({'value': {'$regex': RUN_ID}})
        await db.master_data_values.delete_many({'value': {'$regex': RUN_ID}})
        await db.audit_logs.delete_many({'user_name': {'$regex': RUN_ID}})
        await db.support_tickets.delete_many({'subject': {'$regex': RUN_ID}})
        await db.system_alerts.delete_many({'description': {'$regex': RUN_ID}})
        await db.notification_deliveries.delete_many({'run_id': RUN_ID})
        await db.notification_deliveries.delete_many({'message': {'$regex': RUN_ID}})
        await db.admin_delegations.delete_many({'reason': {'$regex': RUN_ID}})
        await db.org_delegation_requests.delete_many({'org_name': {'$regex': RUN_ID}})
        await db.emergency_requests.delete_many({'reason_text': {'$regex': RUN_ID}})
        await db.emergency_sessions.delete_many({'id': {'$in': _extra_cleanup_ids['sessions']}})
        await db.broadcast_messages.delete_many({'subject': {'$regex': RUN_ID}})
        await db.broadcast_replies.delete_many({'text': {'$regex': RUN_ID}})
        await db.notifications.delete_many({'title': {'$regex': RUN_ID}})
        await db.trials.delete_many({'id': {'$in': _extra_cleanup_ids['trials']}})
        await db.visits.delete_many({'trial_id': {'$in': _extra_cleanup_ids['trials']}})
        await db.patients.delete_many({'email': {'$regex': RUN_ID}})
        await db.visit_instances.delete_many({'trial_id': {'$in': _extra_cleanup_ids['trials']}})
        # reports + their stored CSV blobs
        try:
            import storage as storage_mod
            st = storage_mod.get_storage()
            for _rid, key in _extra_cleanup_ids['reports']:
                await st.delete(key)
        except Exception:
            pass
        await db.admin_reports.delete_many(
            {'id': {'$in': [r for r, _k in _extra_cleanup_ids['reports']]}})
        # remove the test terms versions and restore the seeded active ones;
        # dropping the app_content docs lets /api/legal lazily re-seed defaults
        await db.terms_versions.delete_many(
            {'version': {'$regex': f'^{TERMS_TEST_MAJOR}\\.'}})
        await db.app_content.delete_many({'key': {'$in': ['terms', 'privacy']}})
        for doc_type in ('ToS', 'Privacy'):
            newest_active = await db.terms_versions.find_one(
                {'type': doc_type, 'status': 'active'})
            if not newest_active:
                await db.terms_versions.update_one(
                    {'type': doc_type, 'version': '1.0'},
                    {'$set': {'status': 'active'}})
    run(clean())
    LOOP.close()


@pytest.fixture(scope='module')
def actors():
    async def build():
        admin1, admin1_h = await _make_admin('one')
        admin2, admin2_h = await _make_admin('two')
        pi, pi_h = await _register('pi', org=f'ADMORG-{RUN_ID} Hospital')
        patient, patient_h = await _register('patient')
        return {
            'admin1': (admin1, admin1_h), 'admin2': (admin2, admin2_h),
            'pi': (pi, pi_h), 'patient': (patient, patient_h),
        }
    return run(build())


# ── Role guard: EVERY admin group 403s a non-admin ───────────────────────────
GUARDED_GET_PATHS = [
    '/api/admin/users',
    '/api/admin/users/export',
    '/api/admin/organizations',
    '/api/admin/organizations/duplicates',
    '/api/admin/organizations/name-requests',
    '/api/admin/master-data/submissions',
    '/api/admin/master-data/values',
    '/api/admin/invitations',
    '/api/admin/tickets',
    '/api/admin/alerts',
    '/api/admin/notifications/stats',
    '/api/admin/notifications/log',
    '/api/admin/notifications/settings',
    '/api/admin/audit-logs',
    '/api/admin/audit-logs/summary',
    '/api/admin/audit-logs/security-alerts',
    '/api/admin/audit-logs/export',
    '/api/admin/trials',
    '/api/admin/terms/versions',
    '/api/admin/terms/acceptances',
    '/api/admin/reports/recent',
    '/api/admin/delegations',
    '/api/admin/org-delegation-requests',
    '/api/admin/messages',
    '/api/admin/messages/recipient-count',
    '/api/admin/emergency/requests',
]


class TestAdminRoleGuard:
    def test_non_admin_gets_403_everywhere(self, actors):
        async def flow():
            async with make_client() as cli:
                for path in GUARDED_GET_PATHS:
                    for _, headers in (actors['pi'], actors['patient']):
                        r = await cli.get(path, headers=headers)
                        assert r.status_code == 403, f'{path}: {r.status_code} {r.text}'
        run(flow())

    def test_unauthenticated_gets_401(self):
        async def flow():
            async with make_client() as cli:
                r = await cli.get('/api/admin/users')
                assert r.status_code == 401, r.text
        run(flow())

    def test_admin_passes(self, actors):
        async def flow():
            async with make_client() as cli:
                r = await cli.get('/api/admin/users', headers=actors['admin1'][1])
                assert r.status_code == 200, r.text
                assert isinstance(r.json(), list)
        run(flow())


# ── Users ────────────────────────────────────────────────────────────────────
class TestAdminUsers:
    def test_list_masks_patient_pii(self, actors):
        patient = actors['patient'][0]
        async def flow():
            async with make_client() as cli:
                r = await cli.get('/api/admin/users',
                                  headers=actors['admin1'][1],
                                  params={'search': patient['email']})
                assert r.status_code == 200, r.text
                rows = [u for u in r.json() if u['id'] == patient['id']]
                assert rows, 'patient not found in admin user list'
                masked = rows[0]
                assert masked['full_name'] != patient['full_name']
                assert '***' in masked['full_name']
                assert masked['email'] != patient['email']
                assert patient['full_name'] not in str(masked)
        run(flow())

    def test_create_user_with_invite(self, actors, monkeypatch):
        """Admin user creation NEVER returns a credential — it emails a
        single-use setup link and reports only masked delivery metadata."""
        import otp_service
        sent = []
        monkeypatch.setattr(
            otp_service, 'send_password_reset_email',
            lambda email, link, ttl: sent.append((email, link, ttl)))
        async def flow():
            async with make_client() as cli:
                email = f'adm-{RUN_ID}-created@example.com'
                r = await cli.post('/api/admin/users', headers=actors['admin1'][1], json={
                    'email': email, 'full_name': f'Created User {RUN_ID}',
                    'role': 'crc', 'organization': f'ADMORG-{RUN_ID} Hospital'})
                assert r.status_code == 200, r.text
                j = r.json()
                assert j['user']['status'] == 'Pending Verification'
                assert j['invitation'] and j['invitation']['invite_link']
                # No plaintext credential anywhere in the response.
                assert 'temp_password' not in r.text
                setup = j['password_setup']
                assert setup['reset_sent'] is True
                assert setup['delivery_channel'] == 'email'
                assert setup['email'] != email and '*' in setup['email']
                assert setup['expires_at']
                # The raw link went only to the mocked mailer, never the admin.
                assert len(sent) == 1 and sent[0][0] == email
                assert 'token=' in sent[0][1]
                assert sent[0][1].split('token=')[1] not in r.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.user_create', 'target_id': j['user']['id']})
                assert audit, 'user creation must be audited'
                assert sent[0][1].split('token=')[1] not in str(audit)
        run(flow())

    def test_create_user_rolls_back_on_delivery_failure(self, actors, monkeypatch):
        """If the setup email cannot be delivered, no orphan user, invitation,
        or reset token may remain."""
        import otp_service
        def boom(email, link, ttl):
            raise otp_service.OTPDeliveryError('smtp down')
        monkeypatch.setattr(otp_service, 'send_password_reset_email', boom)
        async def flow():
            async with make_client() as cli:
                email = f'adm-{RUN_ID}-rollback@example.com'
                r = await cli.post('/api/admin/users', headers=actors['admin1'][1], json={
                    'email': email, 'full_name': f'Rollback User {RUN_ID}',
                    'role': 'crc', 'organization': f'ADMORG-{RUN_ID} Hospital'})
                assert r.status_code == 502, r.text
                assert await server.db.users.find_one({'email': email}) is None
                assert await server.db.invitations.find_one({'email': email}) is None
                assert await server.db.password_reset_tokens.find_one({'email': email}) is None
        run(flow())

    def test_suspend_blocks_session_and_login_then_activate(self, actors):
        async def flow():
            async with make_client() as cli:
                victim, victim_h = await _register('crc', org=f'ADMORG-{RUN_ID} Hospital')
                r = await cli.patch(f"/api/admin/users/{victim['id']}/status",
                                    headers=actors['admin1'][1],
                                    json={'status': 'Suspended', 'reason': 'policy violation'})
                assert r.status_code == 200, r.text
                # existing session is dead (403 suspended)
                r2 = await cli.get('/api/auth/me', headers=victim_h)
                assert r2.status_code == 403, r2.text
                # and a fresh login is refused
                r3 = await cli.post('/api/auth/login', json={
                    'email': victim['email'], 'password': PASSWORD})
                assert r3.status_code == 403, r3.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.user_status', 'target_id': victim['id']})
                assert audit, 'status change must be audited'
                # re-activate → login works again
                r4 = await cli.patch(f"/api/admin/users/{victim['id']}/status",
                                     headers=actors['admin1'][1], json={'status': 'Active'})
                assert r4.status_code == 200, r4.text
                r5 = await cli.post('/api/auth/login', json={
                    'email': victim['email'], 'password': PASSWORD})
                assert r5.status_code == 200, r5.text
        run(flow())

    def test_unlock_validations_and_flow(self, actors):
        async def flow():
            async with make_client() as cli:
                victim, _ = await _register('crc', org=f'ADMORG-{RUN_ID} Hospital')
                await server.db.users.update_one({'id': victim['id']}, {'$set': {
                    'status': 'Locked',
                    'lock_info': {'lockedAt': server.now().isoformat(),
                                  'failedAttempts': 5, 'lastIp': '1.2.3.4'}}})
                h = actors['admin1'][1]
                # <2 identity checks → rejected
                r1 = await cli.post(f"/api/admin/users/{victim['id']}/unlock", headers=h,
                                    json={'identity_checks': ['email'],
                                          'reason': 'verified with the user on a call'})
                assert r1.status_code in (400, 422), r1.text
                # reason too short → rejected
                r2 = await cli.post(f"/api/admin/users/{victim['id']}/unlock", headers=h,
                                    json={'identity_checks': ['email', 'phone'],
                                          'reason': 'short'})
                assert r2.status_code == 422, r2.text
                # valid → unlocked
                r3 = await cli.post(f"/api/admin/users/{victim['id']}/unlock", headers=h,
                                    json={'identity_checks': ['email', 'phone'],
                                          'reason': 'verified identity on a recorded call',
                                          'force_password_reset': True})
                assert r3.status_code == 200, r3.text
                fresh = await server.db.users.find_one({'id': victim['id']}, {'_id': 0})
                assert fresh.get('status') == 'Active'
                assert 'lock_info' not in fresh
                assert fresh.get('must_reset_password') is True
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.user_unlock', 'target_id': victim['id']})
                assert audit and audit.get('identity_checks') == ['email', 'phone']
        run(flow())

    def test_reset_password_link_lifecycle_and_force_logout(self, actors, monkeypatch):
        """Admin reset emails a single-use link (no credential in the response),
        the link sets the new password exactly once, and replayed / expired /
        revoked tokens all fail."""
        import hashlib
        import otp_service
        sent = []
        monkeypatch.setattr(
            otp_service, 'send_password_reset_email',
            lambda email, link, ttl: sent.append((email, link, ttl)))
        NEW_PW = f'NewPassw0rd!{RUN_ID}'
        def token_of(link):
            return link.split('token=')[1]
        async def flow():
            async with make_client() as cli:
                victim, victim_h = await _register('crc', org=f'ADMORG-{RUN_ID} Hospital')
                h = actors['admin1'][1]
                r = await cli.post(f"/api/admin/users/{victim['id']}/reset-password", headers=h)
                assert r.status_code == 200, r.text
                j = r.json()
                assert 'temp_password' not in r.text
                assert j['reset_sent'] is True and '*' in j['email']
                assert len(sent) == 1 and sent[0][0] == victim['email']
                token1 = token_of(sent[0][1])
                # existing victim session was force-logged-out by the reset
                dead0 = await cli.get('/api/auth/me', headers=victim_h)
                assert dead0.status_code == 401, dead0.text
                # weak password is refused without consuming the token
                weak = await cli.post('/api/auth/password-reset-link',
                                      json={'token': token1, 'new_password': 'weakpassword'})
                assert weak.status_code == 400, weak.text
                # consume the link with a strong password
                ok = await cli.post('/api/auth/password-reset-link',
                                    json={'token': token1, 'new_password': NEW_PW})
                assert ok.status_code == 200, ok.text
                # old password dead, new password works
                bad = await cli.post('/api/auth/login', json={
                    'email': victim['email'], 'password': PASSWORD})
                assert bad.status_code == 401, bad.text
                good = await cli.post('/api/auth/login', json={
                    'email': victim['email'], 'password': NEW_PW})
                assert good.status_code == 200, good.text
                # replay of the consumed token is refused
                replay = await cli.post('/api/auth/password-reset-link',
                                        json={'token': token1, 'new_password': NEW_PW + 'x'})
                assert replay.status_code == 400, replay.text
                # expired token is refused
                r2 = await cli.post(f"/api/admin/users/{victim['id']}/reset-password", headers=h)
                assert r2.status_code == 200, r2.text
                token2 = token_of(sent[1][1])
                await server.db.password_reset_tokens.update_one(
                    {'token_hash': hashlib.sha256(token2.encode()).hexdigest()},
                    {'$set': {'expires_at': server.now() - timedelta(minutes=1)}})
                expired = await cli.post('/api/auth/password-reset-link',
                                         json={'token': token2, 'new_password': NEW_PW})
                assert expired.status_code == 400, expired.text
                # issuing a newer link revokes the older unused one
                r3 = await cli.post(f"/api/admin/users/{victim['id']}/reset-password", headers=h)
                assert r3.status_code == 200, r3.text
                token3 = token_of(sent[2][1])
                r4 = await cli.post(f"/api/admin/users/{victim['id']}/reset-password", headers=h)
                assert r4.status_code == 200, r4.text
                token4 = token_of(sent[3][1])
                revoked = await cli.post('/api/auth/password-reset-link',
                                         json={'token': token3, 'new_password': NEW_PW})
                assert revoked.status_code == 400, revoked.text
                latest = await cli.post('/api/auth/password-reset-link',
                                        json={'token': token4, 'new_password': NEW_PW})
                assert latest.status_code == 200, latest.text
                # force-logout still invalidates a fresh session token
                good2 = await cli.post('/api/auth/login', json={
                    'email': victim['email'], 'password': NEW_PW})
                assert good2.status_code == 200, good2.text
                fresh_h = {'Authorization': f"Bearer {good2.json()['access_token']}"}
                r5 = await cli.post(f"/api/admin/users/{victim['id']}/force-logout", headers=h)
                assert r5.status_code == 200, r5.text
                dead = await cli.get('/api/auth/me', headers=fresh_h)
                assert dead.status_code == 401, dead.text
        run(flow())

    def test_export_csv(self, actors):
        async def flow():
            async with make_client() as cli:
                r = await cli.get('/api/admin/users/export', headers=actors['admin1'][1])
                assert r.status_code == 200, r.text
                assert 'text/csv' in r.headers['content-type']
                assert r.text.splitlines()[0].startswith('id,name,email')
        run(flow())


# ── Organizations ────────────────────────────────────────────────────────────
class TestAdminOrgs:
    def _mk_org(self, cli_headers, name, otype='site'):
        async def create():
            async with make_client() as cli:
                r = await cli.post('/api/admin/organizations', headers=cli_headers,
                                   json={'name': name, 'type': otype,
                                         'address': 'Test Lane', 'contact': '+91 1'})
                assert r.status_code == 200, r.text
                org = r.json()
                _org_ids.append(org['id'])
                return org
        return run(create())

    def test_create_patch_and_counts(self, actors):
        h = actors['admin1'][1]
        org = self._mk_org(h, f'ADMORG-{RUN_ID} Alpha Clinic')
        async def flow():
            async with make_client() as cli:
                r = await cli.patch(f"/api/admin/organizations/{org['id']}", headers=h,
                                    json={'address': 'New Address 42', 'status': 'suspended'})
                assert r.status_code == 200, r.text
                assert r.json()['address'] == 'New Address 42'
                assert r.json()['status'] == 'suspended'
                lst = await cli.get('/api/admin/organizations', headers=h,
                                    params={'search': f'ADMORG-{RUN_ID} Alpha'})
                assert lst.status_code == 200
                row = [o for o in lst.json() if o['id'] == org['id']][0]
                assert 'users' in row and 'trials' in row
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.org_update', 'target_id': org['id']})
                assert audit, 'org update must be audited'
        run(flow())

    def test_merge_moves_users_and_tombstones_source(self, actors):
        h = actors['admin1'][1]
        source = self._mk_org(h, f'ADMORG-{RUN_ID} Merge Source')
        target = self._mk_org(h, f'ADMORG-{RUN_ID} Merge Target')
        async def flow():
            member, _ = await _register('crc', org=source['name'])
            async with make_client() as cli:
                # justification is mandatory (min 10 chars)
                bad = await cli.post(f"/api/admin/organizations/{source['id']}/merge",
                                     headers=h, json={'target_org_id': target['id'],
                                                      'justification': 'dup'})
                assert bad.status_code == 422, bad.text
                r = await cli.post(f"/api/admin/organizations/{source['id']}/merge",
                                   headers=h, json={
                                       'target_org_id': target['id'],
                                       'justification': 'Duplicate entry for the same hospital'})
                assert r.status_code == 200, r.text
                assert r.json()['moved_users'] >= 1
                moved = await server.db.users.find_one({'id': member['id']}, {'_id': 0})
                assert moved['organization'] == target['name']
                src = await server.db.organizations.find_one({'id': source['id']}, {'_id': 0})
                assert src['status'] == 'merged' and src['merged_into'] == target['id']
                # irreversible: merging again is refused
                again = await cli.post(f"/api/admin/organizations/{source['id']}/merge",
                                       headers=h, json={
                                           'target_org_id': target['id'],
                                           'justification': 'attempting a double merge'})
                assert again.status_code == 400, again.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.org_merge', 'target_id': source['id']})
                assert audit, 'merge must be audited'
        run(flow())

    def test_duplicates_detects_normalized_collisions(self, actors):
        h = actors['admin1'][1]
        self._mk_org(h, f'ADMORG-{RUN_ID} Dup Hospital')
        self._mk_org(h, f'admorg-{RUN_ID} dup  hospital')
        async def flow():
            async with make_client() as cli:
                r = await cli.get('/api/admin/organizations/duplicates', headers=h)
                assert r.status_code == 200, r.text
                names = [o['name'] for g in r.json() for o in g['organizations']]
                assert f'ADMORG-{RUN_ID} Dup Hospital' in names
        run(flow())

    def test_name_request_approve_and_reject(self, actors):
        h = actors['admin1'][1]
        org = self._mk_org(h, f'ADMORG-{RUN_ID} Misspeled Hospitl')
        async def flow():
            db = server.db
            rid1, rid2 = str(uuid.uuid4()), str(uuid.uuid4())
            _extra_cleanup_ids['name_requests'] += [rid1, rid2]
            for rid in (rid1, rid2):
                await db.org_name_requests.insert_one({
                    'id': rid, 'org_id': org['id'], 'current_name': org['name'],
                    'requested_name': f'ADMORG-{RUN_ID} Corrected Hospital',
                    'requested_by': 'someone', 'status': 'pending',
                    'created_at': server.now()})
            async with make_client() as cli:
                final = f'ADMORG-{RUN_ID} Corrected Hospital'
                r = await cli.post(f'/api/admin/organizations/name-requests/{rid1}/approve',
                                   headers=h, json={'finalName': final})
                assert r.status_code == 200, r.text
                fresh = await db.organizations.find_one({'id': org['id']}, {'_id': 0})
                assert fresh['name'] == final
                r2 = await cli.post(f'/api/admin/organizations/name-requests/{rid2}/reject',
                                    headers=h, json={'reason': 'Name already corrected'})
                assert r2.status_code == 200, r2.text
                req2 = await db.org_name_requests.find_one({'id': rid2}, {'_id': 0})
                assert req2['status'] == 'rejected'
        run(flow())


# ── Master data ──────────────────────────────────────────────────────────────
class TestAdminMasterData:
    async def _mk_submission(self, value, field_type='designation', submitted_by_id=None):
        sid = str(uuid.uuid4())
        _extra_cleanup_ids['submissions'].append(sid)
        doc = {
            'id': sid, 'fieldType': field_type, 'value': value,
            'submittedBy': f'Test {RUN_ID}', 'org': f'ADMORG-{RUN_ID} Hospital',
            'dateSubmitted': server.now(), 'status': 'pending',
            'actionBy': None, 'rejectReason': ''}
        if submitted_by_id:
            doc['submittedById'] = submitted_by_id
        await server.db.master_data_submissions.insert_one(doc)
        return sid

    def test_approve_adds_global_value(self, actors):
        h = actors['admin1'][1]
        async def flow():
            sid = await self._mk_submission(f'Trial Coordinator {RUN_ID}')
            async with make_client() as cli:
                r = await cli.post(f'/api/admin/master-data/submissions/{sid}/approve',
                                   headers=h, json={})
                assert r.status_code == 200, r.text
                vals = await cli.get('/api/admin/master-data/values', headers=h,
                                     params={'fieldType': 'designation'})
                assert any(v['value'] == f'Trial Coordinator {RUN_ID}' for v in vals.json())
                # double-approve refused
                again = await cli.post(f'/api/admin/master-data/submissions/{sid}/approve',
                                       headers=h, json={})
                assert again.status_code == 400, again.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.master_data_approve', 'target_id': sid})
                assert audit, 'approve must be audited'
        run(flow())

    def test_edit_and_approve_uses_corrected_value(self, actors):
        h = actors['admin1'][1]
        async def flow():
            sid = await self._mk_submission(f'reserch fellow {RUN_ID}')
            corrected = f'Research Fellow {RUN_ID}'
            async with make_client() as cli:
                r = await cli.post(f'/api/admin/master-data/submissions/{sid}/approve',
                                   headers=h, json={'value': corrected})
                assert r.status_code == 200, r.text
                assert r.json()['value'] == corrected
                sub = await server.db.master_data_submissions.find_one({'id': sid}, {'_id': 0})
                assert sub['value'] == corrected and sub['status'] == 'approved'
        run(flow())

    def test_reject_requires_reason(self, actors):
        h = actors['admin1'][1]
        async def flow():
            sid = await self._mk_submission(f'Wellness Guru {RUN_ID}')
            async with make_client() as cli:
                bad = await cli.post(f'/api/admin/master-data/submissions/{sid}/reject',
                                     headers=h, json={'reason': ''})
                assert bad.status_code == 422, bad.text
                r = await cli.post(f'/api/admin/master-data/submissions/{sid}/reject',
                                   headers=h, json={'reason': 'Not a clinical designation'})
                assert r.status_code == 200, r.text
                sub = await server.db.master_data_submissions.find_one({'id': sid}, {'_id': 0})
                assert sub['status'] == 'rejected'
                assert sub['rejectReason'] == 'Not a clinical designation'
        run(flow())

    def test_department_visibility_and_edit_approval(self, actors):
        admin_h = actors['admin1'][1]
        pi, pi_h = actors['pi']
        original = f'Experimental Department {RUN_ID}'
        corrected = f'Experimental Medicine {RUN_ID}'

        async def flow():
            sid = await self._mk_submission(
                original, field_type='department', submitted_by_id=pi['id'])
            await server.db.users.update_one({'id': pi['id']}, {'$set': {
                'profile.department': original,
                'profile.department_review_status': 'pending',
            }})
            async with make_client() as cli:
                public = await cli.get('/api/master-data/options', params={
                    'fieldType': 'department'})
                private = await cli.get('/api/master-data/options', headers=pi_h,
                                        params={'fieldType': 'department'})
                assert original not in public.json()['values']
                assert any(row['value'] == original
                           for row in private.json()['private_values'])

                approved = await cli.post(
                    f'/api/admin/master-data/submissions/{sid}/approve',
                    headers=admin_h, json={'value': corrected})
                assert approved.status_code == 200, approved.text
                published = await cli.get('/api/master-data/options', params={
                    'fieldType': 'department'})
                assert corrected in published.json()['values']

            fresh = await server.db.users.find_one({'id': pi['id']}, {'_id': 0})
            assert fresh['profile']['department'] == corrected
            assert fresh['profile']['department_review_status'] == 'approved'

            rejected_value = f'Private Department {RUN_ID}'
            rejected_sid = await self._mk_submission(
                rejected_value, field_type='department', submitted_by_id=pi['id'])
            await server.db.users.update_one({'id': pi['id']}, {'$set': {
                'profile.department': rejected_value,
                'profile.department_review_status': 'pending',
            }})
            async with make_client() as cli:
                rejected = await cli.post(
                    f'/api/admin/master-data/submissions/{rejected_sid}/reject',
                    headers=admin_h, json={'reason': 'Too site-specific'})
                assert rejected.status_code == 200, rejected.text
                public = await cli.get('/api/master-data/options', params={
                    'fieldType': 'department'})
                private = await cli.get('/api/master-data/options', headers=pi_h,
                                        params={'fieldType': 'department'})
                assert rejected_value not in public.json()['values']
                assert any(row['value'] == rejected_value and row['status'] == 'rejected'
                           for row in private.json()['private_values'])

            fresh = await server.db.users.find_one({'id': pi['id']}, {'_id': 0})
            assert fresh['profile']['department'] == rejected_value
            assert fresh['profile']['department_review_status'] == 'rejected'

        run(flow())


# ── Tickets (admin triage) ───────────────────────────────────────────────────
class TestAdminTickets:
    def test_triage_note_and_status(self, actors):
        h = actors['admin1'][1]
        pi, pi_h = actors['pi']
        async def flow():
            async with make_client() as cli:
                # user files a ticket through the untouched user-side endpoint
                r = await cli.post('/api/support/tickets', headers=pi_h, json={
                    'category': 'Technical',
                    'subject': f'Cannot open schedule {RUN_ID}',
                    'description': 'The schedule tab crashes.'})
                assert r.status_code == 200, r.text
                ticket = r.json()
                # admin sees it with the reporter enriched
                lst = await cli.get('/api/admin/tickets', headers=h,
                                    params={'search': RUN_ID})
                assert lst.status_code == 200, lst.text
                mine = [t for t in lst.json() if t['id'] == ticket['id']]
                assert mine and mine[0]['user']['id'] == pi['id']
                # add a note → stored + owner notified
                rn = await cli.post(f"/api/admin/tickets/{ticket['id']}/notes", headers=h,
                                    json={'text': f'Looking into this {RUN_ID}'})
                assert rn.status_code == 200, rn.text
                # change status + priority
                rp = await cli.patch(f"/api/admin/tickets/{ticket['id']}", headers=h,
                                     json={'status': 'In Progress', 'priority': 'high'})
                assert rp.status_code == 200, rp.text
                assert rp.json()['status'] == 'In Progress'
                fresh = await server.db.support_tickets.find_one({'id': ticket['id']}, {'_id': 0})
                assert fresh['notes'][0]['text'] == f'Looking into this {RUN_ID}'
                assert fresh['priority'] == 'high'
                notif = await server.db.notifications.find_one(
                    {'user_id': pi['id'], 'kind': 'support'})
                assert notif, 'ticket owner must be notified'
                for action in ('admin.ticket_note', 'admin.ticket_update'):
                    audit = await server.db.audit_logs.find_one(
                        {'action': action, 'target_id': ticket['id']})
                    assert audit, f'{action} must be audited'
        run(flow())


# ── System alerts ────────────────────────────────────────────────────────────
class TestAdminAlerts:
    def test_alert_lifecycle(self, actors):
        h = actors['admin1'][1]
        patient = actors['patient'][0]
        async def flow():
            aid = str(uuid.uuid4())
            await server.db.system_alerts.insert_one({
                'id': aid, 'type': 'OTP failure',
                'description': f'OTP failed for run {RUN_ID}',
                'affected': patient['email'], 'severity': 'high',
                'status': 'open', 'timestamp': server.now()})
            async with make_client() as cli:
                r0 = await cli.get('/api/admin/alerts', headers=h, params={'status': 'open'})
                assert any(a['id'] == aid for a in r0.json())
                r1 = await cli.post(f'/api/admin/alerts/{aid}/retry', headers=h)
                assert r1.status_code == 200, r1.text
                r2 = await cli.post(f'/api/admin/alerts/{aid}/notify-user', headers=h)
                assert r2.status_code == 200, r2.text
                r3 = await cli.post(f'/api/admin/alerts/{aid}/escalate', headers=h)
                assert r3.status_code == 200 and r3.json()['severity'] == 'critical'
                r4 = await cli.post(f'/api/admin/alerts/{aid}/resolve', headers=h,
                                    json={'note': 'SMS provider restored'})
                assert r4.status_code == 200, r4.text
                fresh = await server.db.system_alerts.find_one({'id': aid}, {'_id': 0})
                assert fresh['status'] == 'resolved'
                assert fresh['resolution_note'] == 'SMS provider restored'
                # resolving twice is refused
                again = await cli.post(f'/api/admin/alerts/{aid}/resolve', headers=h, json={})
                assert again.status_code == 400, again.text
                for action in ('admin.alert_retry', 'admin.alert_notify_user',
                               'admin.alert_escalate', 'admin.alert_resolve'):
                    audit = await server.db.audit_logs.find_one(
                        {'action': action, 'target_id': aid})
                    assert audit, f'{action} must be audited'
        run(flow())


# ── Notification monitoring ──────────────────────────────────────────────────
class TestAdminNotifMonitoring:
    def test_log_masks_recipient_and_retry(self, actors):
        h = actors['admin1'][1]
        async def flow():
            did = str(uuid.uuid4())
            await server.db.notification_deliveries.insert_one({
                'id': did, 'run_id': RUN_ID, 'type': 'visit_reminder', 'channel': 'SMS',
                'recipient': '+91 9876543210', 'message': f'Reminder {RUN_ID}',
                'status': 'Failed', 'sentAt': server.now(), 'error': 'carrier timeout'})
            async with make_client() as cli:
                log = await cli.get('/api/admin/notifications/log', headers=h,
                                    params={'status': 'Failed'})
                mine = [d for d in log.json() if d['id'] == did]
                assert mine, 'delivery row missing from log'
                assert '+91 9876543210' not in str(mine[0]), 'recipient must be masked'
                assert '*' in mine[0]['recipient']
                stats = await cli.get('/api/admin/notifications/stats', headers=h)
                assert stats.status_code == 200
                assert stats.json()['by_status']['failed'] >= 1
                r = await cli.post(f'/api/admin/notifications/{did}/retry', headers=h)
                assert r.status_code == 200, r.text
                fresh = await server.db.notification_deliveries.find_one({'id': did}, {'_id': 0})
                assert fresh['status'] == 'Pending'
                # retrying a non-failed delivery is refused
                again = await cli.post(f'/api/admin/notifications/{did}/retry', headers=h)
                assert again.status_code == 400, again.text
        run(flow())

    def test_settings_roundtrip(self, actors):
        h = actors['admin1'][1]
        async def flow():
            async with make_client() as cli:
                r0 = await cli.get('/api/admin/notifications/settings', headers=h)
                assert r0.status_code == 200
                assert 'visitReminderHours' in r0.json()
                r1 = await cli.patch('/api/admin/notifications/settings', headers=h,
                                     json={'visitReminderHours': 48,
                                           'channels': {'sms': False}})
                assert r1.status_code == 200, r1.text
                j = r1.json()
                assert j['visitReminderHours'] == 48
                assert j['channels']['sms'] is False
                assert j['channels']['push'] is True   # merged, not replaced
                bad = await cli.patch('/api/admin/notifications/settings', headers=h,
                                      json={'channels': {'pigeon': True}})
                assert bad.status_code == 400, bad.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.notification_settings'})
                assert audit, 'settings change must be audited'
        run(flow())


# ── Audit views ──────────────────────────────────────────────────────────────
class TestAdminAudit:
    def test_admin_scope_summary_alerts_export(self, actors):
        h = actors['admin1'][1]
        admin1 = actors['admin1'][0]
        async def flow():
            # seed a burst of failed logins for the security-alert pattern
            for _ in range(3):
                await server.write_audit(
                    {'id': f'victim-{RUN_ID}', 'full_name': f'Victim {RUN_ID}',
                     'role': 'crc', 'organization': ''},
                    'login.failed', 'Wrong password', status='failure', ip='9.9.9.9')
            async with make_client() as cli:
                r = await cli.get('/api/admin/audit-logs', headers=h,
                                  params={'category': 'login', 'status': 'failure'})
                assert r.status_code == 200 and len(r.json()) >= 3
                s = await cli.get('/api/admin/audit-logs/summary', headers=h)
                assert s.status_code == 200
                assert s.json()['total'] >= 3 and 'by_category' in s.json()
                a = await cli.get('/api/admin/audit-logs/security-alerts', headers=h)
                assert a.status_code == 200
                assert any(x['user_id'] == f'victim-{RUN_ID}' and x['count'] >= 3
                           for x in a.json()), 'failed-login pattern not detected'
                e = await cli.get('/api/admin/audit-logs/export', headers=h,
                                  params={'category': 'login'})
                assert e.status_code == 200 and 'text/csv' in e.headers['content-type']
                assert e.text.splitlines()[0].startswith('time,user')
        run(flow())


# ── Admin trials (aggregates only; masked subjects) ──────────────────────────
@pytest.fixture(scope='module')
def trial_world(actors):
    async def build():
        sponsor, sponsor_h = await _register('sponsor', org=f'ADMORG-{RUN_ID} Pharma')
        async with make_client() as cli:
            r = await cli.post('/api/trials', headers=sponsor_h, json={
                'title': f'Admin Trial {RUN_ID}', 'protocol_id': f'ADM-{RUN_ID}',
                'phase': 'Phase II', 'condition': 'Testing',
                'sponsor_name': f'ADMORG-{RUN_ID} Pharma'})
            assert r.status_code == 200, r.text
            trial = r.json()
            _extra_cleanup_ids['trials'].append(trial['id'])
            rv = await cli.post('/api/visits', headers=sponsor_h, json={
                'trial_id': trial['id'], 'visit_number': 1, 'name': 'Screening',
                'day_offset': 0, 'window_days': 3, 'activities': ['Vitals']})
            assert rv.status_code == 200, rv.text
            pi, pi_h = await _register('pi', org=f'ADMORG-{RUN_ID} TrialSite')
            await server.db.invitations.insert_one({
                'id': str(uuid.uuid4()), 'code': f'ADM-TRIAL-{RUN_ID}',
                'email': pi['email'], 'phone': pi.get('phone', ''),
                'role': 'pi', 'trial_id': trial['id'],
                'status': 'accepted', 'created_at': server.now(),
            })
            async with make_client() as cli:
                rp = await cli.post('/api/patients', headers=pi_h, json={
                    'full_name': f'Secret Patient {RUN_ID}',
                    'email': f'adm-{RUN_ID}-subject@example.com',
                    'trial_id': trial['id'], 'pi_id': pi['id'],
                    'enrolled_date': (server.now() - timedelta(days=3)).date().isoformat()})
                assert rp.status_code == 200, rp.text
                patient = rp.json()
        return {'trial': trial, 'patient': patient, 'sponsor_h': sponsor_h}
    return run(build())


class TestAdminTrials:
    def test_list_and_detail_are_masked(self, actors, trial_world):
        h = actors['admin1'][1]
        sponsor_h = trial_world['sponsor_h']
        trial = trial_world['trial']
        patient = trial_world['patient']
        async def flow():
            async with make_client() as cli:
                changed = await cli.patch(
                    f"/api/trials/{trial['id']}", headers=sponsor_h,
                    json={'target_enrollment': 25})
                assert changed.status_code == 200, changed.text
                lst = await cli.get('/api/admin/trials', headers=h)
                assert lst.status_code == 200, lst.text
                mine = [t for t in lst.json() if t['id'] == trial['id']]
                assert mine and mine[0]['patients'] == 1
                assert mine[0]['sponsorOrCroName'] == f'ADMORG-{RUN_ID} Pharma'
                assert mine[0]['ownerType'] == 'Sponsor'
                assert mine[0]['modifiedBy']
                assert mine[0]['changeSummary'] == 'Updated target enrollment'
                assert mine[0]['changedFields'] == ['target_enrollment']
                assert f'Secret Patient {RUN_ID}' not in lst.text, 'list must carry no PII'
                det = await cli.get(f"/api/admin/trials/{trial['id']}", headers=h)
                assert det.status_code == 200, det.text
                j = det.json()
                assert j['unmasked'] is False
                assert j['subjects'] and j['subjects'][0]['subject'].startswith('SUBJ-')
                body = det.text
                assert patient['full_name'] not in body
                assert patient['email'] not in body
        run(flow())


# ── Admin invitations ────────────────────────────────────────────────────────
class TestAdminInvitations:
    def test_create_resend_cancel(self, actors):
        h = actors['admin1'][1]
        async def flow():
            async with make_client() as cli:
                r = await cli.post('/api/admin/invitations', headers=h, json={
                    'email': f'adm-{RUN_ID}-invitee@example.com',
                    'full_name': f'Invitee {RUN_ID}', 'role': 'crc',
                    'organization': f'ADMORG-{RUN_ID} Hospital'})
                assert r.status_code == 200, r.text
                inv = r.json()
                assert inv['invite_link']
                lst = await cli.get('/api/admin/invitations', headers=h)
                assert any(i['id'] == inv['id'] for i in lst.json())
                r2 = await cli.post(f"/api/admin/invitations/{inv['id']}/resend", headers=h)
                assert r2.status_code == 200, r2.text
                r3 = await cli.post(f"/api/admin/invitations/{inv['id']}/cancel", headers=h)
                assert r3.status_code == 200, r3.text
                fresh = await server.db.invitations.find_one({'id': inv['id']}, {'_id': 0})
                assert fresh['status'] == 'cancelled'
                for action in ('admin.invitation_create', 'admin.invitation_resend',
                               'admin.invitation_cancel'):
                    audit = await server.db.audit_logs.find_one(
                        {'action': action, 'target_id': inv['id']})
                    assert audit, f'{action} must be audited'
        run(flow())


# ── Terms & privacy versions ─────────────────────────────────────────────────
class TestAdminTerms:
    def test_publish_supersede_and_force_reacceptance(self, actors):
        h = actors['admin1'][1]
        patient, patient_h = actors['patient']
        v1 = f'{TERMS_TEST_MAJOR}.0'
        async def flow():
            async with make_client() as cli:
                # the patient accepts the current terms first
                ra = await cli.post('/api/legal/accept', headers=patient_h)
                assert ra.status_code == 200, ra.text
                acc = await cli.get('/api/admin/terms/acceptances', headers=h)
                assert any(a['user_id'] == patient['id'] for a in acc.json())
                # non-numeric version refused
                bad = await cli.post('/api/admin/terms/versions', headers=h, json={
                    'type': 'ToS', 'version': 'abc', 'content': 'x'})
                assert bad.status_code == 400, bad.text
                # publish with forced re-acceptance
                r = await cli.post('/api/admin/terms/versions', headers=h, json={
                    'type': 'ToS', 'version': v1,
                    'content': f'Updated terms for run {RUN_ID}.',
                    'changeSummary': f'Test update {RUN_ID}',
                    'forceReacceptance': True})
                assert r.status_code == 200, r.text
                j = r.json()
                assert j['status'] == 'active'
                assert j['reacceptance_required'] >= 1
                # the previous active ToS is superseded; only one active remains
                actives = await server.db.terms_versions.count_documents(
                    {'type': 'ToS', 'status': 'active'})
                assert actives == 1
                # the patient's acceptance was cleared → must re-accept
                fresh = await server.db.users.find_one({'id': patient['id']}, {'_id': 0})
                assert 'terms_accepted_at' not in fresh
                # same version again → 400 (must be strictly greater)
                dup = await cli.post('/api/admin/terms/versions', headers=h, json={
                    'type': 'ToS', 'version': v1, 'content': 'y'})
                assert dup.status_code == 400, dup.text
                # user-facing /api/legal now serves the published content
                legal = await cli.get('/api/legal/terms')
                assert legal.status_code == 200
                assert legal.json()['version'] == v1
                # edit content on the active version
                pe = await cli.patch(f"/api/admin/terms/versions/{j['id']}", headers=h,
                                     json={'content': f'Amended terms for run {RUN_ID}.'})
                assert pe.status_code == 200, pe.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.terms_publish', 'target_id': j['id']})
                assert audit and audit.get('forceReacceptance') is True
        run(flow())


# ── Reports ──────────────────────────────────────────────────────────────────
class TestAdminReports:
    def test_generate_recent_download(self, actors):
        h = actors['admin1'][1]
        async def flow():
            async with make_client() as cli:
                r = await cli.post('/api/admin/reports/generate', headers=h,
                                   json={'type': 'user-status', 'format': 'pdf'})
                assert r.status_code == 200, r.text
                rep = r.json()
                _extra_cleanup_ids['reports'].append((rep['id'], rep['key']))
                assert rep['name'].startswith('user-status-')
                recent = await cli.get('/api/admin/reports/recent', headers=h)
                assert any(x['id'] == rep['id'] for x in recent.json())
                dl = await cli.get(f"/api/admin/reports/{rep['id']}/download", headers=h)
                assert dl.status_code == 200, dl.text
                assert 'application/pdf' in dl.headers['content-type']
                assert dl.content.startswith(b'%PDF')
                # patient PII never appears in a users report
                r2 = await cli.post('/api/admin/reports/generate', headers=h,
                                    json={'type': 'users', 'format': 'xlsx'})
                rep2 = r2.json()
                _extra_cleanup_ids['reports'].append((rep2['id'], rep2['key']))
                dl2 = await cli.get(f"/api/admin/reports/{rep2['id']}/download", headers=h)
                assert 'spreadsheetml' in dl2.headers['content-type']
                assert dl2.content.startswith(b'PK')
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(dl2.content), read_only=True)
                values = '\n'.join(
                    str(value or '')
                    for row in workbook.active.iter_rows(values_only=True)
                    for value in row
                )
                patient = actors['patient'][0]
                assert patient['full_name'] not in values
                assert patient['email'] not in values
                # unknown report type is rejected by the schema
                bad = await cli.post('/api/admin/reports/generate', headers=h,
                                     json={'type': 'everything'})
                assert bad.status_code == 422, bad.text
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.report_generate', 'target_id': rep['id']})
                assert audit, 'report generation must be audited'
        run(flow())


# ── Delegations ──────────────────────────────────────────────────────────────
class TestAdminDelegations:
    def test_crud_lifecycle(self, actors):
        h = actors['admin1'][1]
        pi = actors['pi'][0]
        patient = actors['patient'][0]
        reason = f'Backup coverage while primary admin is on leave {RUN_ID}'
        async def flow():
            async with make_client() as cli:
                # reason < 20 chars refused
                bad = await cli.post('/api/admin/delegations', headers=h, json={
                    'user_id': pi['id'], 'tasks': ['support_tickets'], 'reason': 'short'})
                assert bad.status_code == 422, bad.text
                # a patient can never hold admin delegation
                badp = await cli.post('/api/admin/delegations', headers=h, json={
                    'user_id': patient['id'], 'tasks': ['support_tickets'],
                    'reason': reason})
                assert badp.status_code == 400, badp.text
                r = await cli.post('/api/admin/delegations', headers=h, json={
                    'user_id': pi['id'],
                    'tasks': ['support_tickets', 'invitations'], 'reason': reason})
                assert r.status_code == 200, r.text
                d = r.json()
                # duplicate active delegation refused
                dup = await cli.post('/api/admin/delegations', headers=h, json={
                    'user_id': pi['id'], 'tasks': ['reports'], 'reason': reason})
                assert dup.status_code == 400, dup.text
                # edit tasks
                pe = await cli.patch(f"/api/admin/delegations/{d['id']}", headers=h,
                                     json={'tasks': ['reports', 'audit_review']})
                assert pe.status_code == 200, pe.text
                assert sorted(pe.json()['tasks']) == ['audit_review', 'reports']
                # suspend then revoke; revoked cannot be edited
                s = await cli.post(f"/api/admin/delegations/{d['id']}/suspend", headers=h)
                assert s.status_code == 200, s.text
                rv = await cli.delete(f"/api/admin/delegations/{d['id']}", headers=h)
                assert rv.status_code == 200 and rv.json()['status'] == 'revoked'
                fresh = await server.db.admin_delegations.find_one({'id': d['id']}, {'_id': 0})
                assert fresh['status'] == 'revoked', 'record must be kept, not deleted'
                pe2 = await cli.patch(f"/api/admin/delegations/{d['id']}", headers=h,
                                      json={'tasks': ['reports']})
                assert pe2.status_code == 400, pe2.text
                for action in ('admin.delegation_create', 'admin.delegation_update',
                               'admin.delegation_suspend', 'admin.delegation_revoke'):
                    audit = await server.db.audit_logs.find_one(
                        {'action': action, 'target_id': d['id']})
                    assert audit, f'{action} must be audited'
        run(flow())

    def test_org_trial_creation_request_approve_and_reject(self, actors):
        h = actors['admin1'][1]
        requester = actors['pi'][0]

        async def flow():
            async with make_client() as cli:
                approved_org = {
                    'id': str(uuid.uuid4()),
                    'name': f'ADM-DELEGATE-SITE-{RUN_ID}',
                    'type': 'site',
                    'trial_creation_delegated': False,
                    'created_at': server.now(),
                }
                rejected_org = {
                    'id': str(uuid.uuid4()),
                    'name': f'ADM-DELEGATE-SMO-{RUN_ID}',
                    'type': 'smo',
                    'trial_creation_delegated': False,
                    'created_at': server.now(),
                }
                await server.db.organizations.insert_many([approved_org, rejected_org])
                requests = [{
                    'id': str(uuid.uuid4()), 'org_id': approved_org['id'],
                    'org_name': approved_org['name'], 'requested_by': requester['id'],
                    'requester_name': requester['full_name'],
                    'reason': f'Site needs delegated trial creation {RUN_ID}',
                    'status': 'pending', 'created_at': server.now(),
                }, {
                    'id': str(uuid.uuid4()), 'org_id': rejected_org['id'],
                    'org_name': rejected_org['name'], 'requested_by': requester['id'],
                    'requester_name': requester['full_name'],
                    'reason': f'SMO request for review {RUN_ID}',
                    'status': 'pending', 'created_at': server.now(),
                }]
                await server.db.org_delegation_requests.insert_many(requests)

                listed = await cli.get(
                    '/api/admin/org-delegation-requests?status=pending', headers=h)
                assert listed.status_code == 200, listed.text
                ids = {row['id'] for row in listed.json()}
                assert requests[0]['id'] in ids and requests[1]['id'] in ids

                approved = await cli.post(
                    f"/api/admin/org-delegation-requests/{requests[0]['id']}/approve",
                    headers=h, json={'reason': 'Verified operating controls'})
                assert approved.status_code == 200, approved.text
                assert approved.json()['status'] == 'approved'
                approved_fresh = await server.db.organizations.find_one(
                    {'id': approved_org['id']}, {'_id': 0})
                assert approved_fresh['trial_creation_delegated'] is True

                duplicate = await cli.post(
                    f"/api/admin/org-delegation-requests/{requests[0]['id']}/reject",
                    headers=h, json={'reason': 'Too late'})
                assert duplicate.status_code == 400, duplicate.text

                rejected = await cli.post(
                    f"/api/admin/org-delegation-requests/{requests[1]['id']}/reject",
                    headers=h, json={'reason': 'Required controls are incomplete'})
                assert rejected.status_code == 200, rejected.text
                rejected_fresh = await server.db.organizations.find_one(
                    {'id': rejected_org['id']}, {'_id': 0})
                assert rejected_fresh['trial_creation_delegated'] is False

                for action, request in (
                    ('admin.org_delegation_approved', requests[0]),
                    ('admin.org_delegation_rejected', requests[1]),
                ):
                    audit = await server.db.audit_logs.find_one({
                        'action': action, 'target_id': request['id']})
                    assert audit, f'{action} must be audited'
                note = await server.db.notifications.find_one({
                    'user_id': requester['id'],
                    'title': {'$regex': 'Trial creation request'}})
                assert note, 'requester must be notified of the decision'
        run(flow())


# ── Emergency access (break-the-glass) ───────────────────────────────────────
class TestEmergencyAccess:
    def test_full_btg_lifecycle(self, actors, trial_world):
        admin1, h1 = actors['admin1']
        admin2, h2 = actors['admin2']
        trial = trial_world['trial']
        patient = trial_world['patient']
        async def flow():
            async with make_client() as cli:
                # non-admin cannot even request
                guard = await cli.post('/api/admin/emergency/requests',
                                       headers=actors['pi'][1],
                                       json={'reason_category': 'patient_safety',
                                             'reason_text': f'guard check {RUN_ID}'})
                assert guard.status_code == 403, guard.text
                # request
                r = await cli.post('/api/admin/emergency/requests', headers=h1, json={
                    'reason_category': 'patient_safety',
                    'reason_text': f'Urgent safety signal review {RUN_ID}',
                    'trial_id': trial['id']})
                assert r.status_code == 200, r.text
                req = r.json()
                inbox = await cli.get(
                    '/api/admin/emergency/requests?status=pending', headers=h2)
                assert inbox.status_code == 200, inbox.text
                listed = [row for row in inbox.json() if row['id'] == req['id']]
                assert listed and listed[0]['can_action'] is True
                own_inbox = await cli.get(
                    '/api/admin/emergency/requests?status=pending', headers=h1)
                own = [row for row in own_inbox.json() if row['id'] == req['id']]
                assert own and own[0]['can_action'] is False
                # requester cannot self-approve (two-person rule)
                self_ok = await cli.post(
                    f"/api/admin/emergency/requests/{req['id']}/approve", headers=h1)
                assert self_ok.status_code == 403, self_ok.text
                # second admin approves → 2h session
                ap = await cli.post(
                    f"/api/admin/emergency/requests/{req['id']}/approve", headers=h2)
                assert ap.status_code == 200, ap.text
                session = ap.json()['session']
                _extra_cleanup_ids['sessions'].append(session['id'])
                started = server.datetime.fromisoformat(session['started_at'])
                expires = server.datetime.fromisoformat(session['expires_at'])
                assert abs((expires - started).total_seconds() - 7200) < 5, \
                    'session TTL must be 2 hours'
                # poll shows the active session
                poll = await cli.get(f"/api/admin/emergency/requests/{req['id']}", headers=h1)
                assert poll.json()['session']['status'] == 'active'
                # during the session the requester sees IDENTIFIED subjects…
                det = await cli.get(f"/api/admin/trials/{trial['id']}", headers=h1)
                assert det.json()['unmasked'] is True
                assert any(s.get('full_name') == patient['full_name']
                           for s in det.json()['subjects'])
                # …and that read is in the session audit log
                log = await cli.get(
                    f"/api/admin/emergency/sessions/{session['id']}/log", headers=h2)
                assert any(e['action'] == 'emergency.read' for e in log.json()), \
                    'unmasked reads must be audited with the session id'
                # the approver has NO session → still masked
                det2 = await cli.get(f"/api/admin/trials/{trial['id']}", headers=h2)
                assert det2.json()['unmasked'] is False
                # end the session → reads are masked again
                end = await cli.post(
                    f"/api/admin/emergency/sessions/{session['id']}/end", headers=h1)
                assert end.status_code == 200, end.text
                det3 = await cli.get(f"/api/admin/trials/{trial['id']}", headers=h1)
                assert det3.json()['unmasked'] is False
        run(flow())

    def test_expiry_is_enforced(self, actors, trial_world):
        admin1, h1 = actors['admin1']
        h2 = actors['admin2'][1]
        trial = trial_world['trial']
        async def flow():
            async with make_client() as cli:
                r = await cli.post('/api/admin/emergency/requests', headers=h1, json={
                    'reason_category': 'incident_investigation',
                    'reason_text': f'Expiry enforcement check {RUN_ID}'})
                req = r.json()
                ap = await cli.post(
                    f"/api/admin/emergency/requests/{req['id']}/approve", headers=h2)
                session = ap.json()['session']
                _extra_cleanup_ids['sessions'].append(session['id'])
                # simulate the 2h TTL elapsing
                await server.db.emergency_sessions.update_one(
                    {'id': session['id']},
                    {'$set': {'expires_at': server.now() - timedelta(minutes=1)}})
                det = await cli.get(f"/api/admin/trials/{trial['id']}", headers=h1)
                assert det.json()['unmasked'] is False, 'expired session must not unmask'
                fresh = await server.db.emergency_sessions.find_one(
                    {'id': session['id']}, {'_id': 0})
                assert fresh['status'] == 'expired'
                # ending an expired session is refused
                end = await cli.post(
                    f"/api/admin/emergency/sessions/{session['id']}/end", headers=h1)
                assert end.status_code == 400, end.text
        run(flow())

    def test_deny_flow(self, actors):
        h1 = actors['admin1'][1]
        h2 = actors['admin2'][1]
        async def flow():
            async with make_client() as cli:
                r = await cli.post('/api/admin/emergency/requests', headers=h1, json={
                    'reason_category': 'other',
                    'reason_text': f'Deny flow check {RUN_ID}'})
                req = r.json()
                d = await cli.post(f"/api/admin/emergency/requests/{req['id']}/deny",
                                   headers=h2, json={'reason': 'No justification'})
                assert d.status_code == 200, d.text
                # approving a denied request is refused
                ap = await cli.post(
                    f"/api/admin/emergency/requests/{req['id']}/approve", headers=h2)
                assert ap.status_code == 400, ap.text
        run(flow())


# ── Broadcast messages ───────────────────────────────────────────────────────
class TestAdminMessages:
    def test_recipient_count_does_not_send(self, actors):
        h = actors['admin1'][1]
        async def flow():
            db = server.db
            before = await db.notifications.count_documents({'kind': 'broadcast'})
            async with make_client() as cli:
                r = await cli.get('/api/admin/messages/recipient-count', headers=h,
                                  params={'target': 'all'})
                assert r.status_code == 200 and r.json()['count'] >= 4
                r2 = await cli.get('/api/admin/messages/recipient-count', headers=h,
                                   params={'target': 'role:pi'})
                assert r2.status_code == 200 and r2.json()['count'] >= 1
                bad = await cli.get('/api/admin/messages/recipient-count', headers=h,
                                    params={'target': 'galaxy:andromeda'})
                assert bad.status_code == 400, bad.text
            after = await db.notifications.count_documents({'kind': 'broadcast'})
            assert after == before, 'recipient-count must never fan out'
        run(flow())

    def test_broadcast_fans_out_and_tracks_reads(self, actors):
        h = actors['admin1'][1]
        patient = actors['patient'][0]
        subject = f'Please update the app {RUN_ID}'
        async def flow():
            async with make_client() as cli:
                # subject > 120 chars refused
                bad = await cli.post('/api/admin/messages', headers=h, json={
                    'subject': 'x' * 121, 'body': 'b', 'target': f"user:{patient['id']}"})
                assert bad.status_code == 422, bad.text
                r = await cli.post('/api/admin/messages', headers=h, json={
                    'type': 'targeted', 'subject': subject,
                    'body': 'A new version is available.',
                    'target': f"user:{patient['id']}"})
                assert r.status_code == 200, r.text
                b = r.json()
                assert b['status'] == 'sent' and b['recipients_count'] == 1
                notif = await server.db.notifications.find_one(
                    {'broadcast_id': b['id'], 'user_id': patient['id']}, {'_id': 0})
                assert notif, 'broadcast must fan out to the notifications collection'
                delivery = await server.db.notification_deliveries.find_one(
                    {'broadcast_id': b['id']}, {'_id': 0})
                assert delivery and delivery['status'] == 'Delivered'
                # recipient reads it → read x/y reflects it
                await server.db.notifications.update_one(
                    {'id': notif['id']}, {'$set': {'read': True}})
                lst = await cli.get('/api/admin/messages', headers=h,
                                    params={'box': 'sent'})
                mine = [m for m in lst.json() if m['id'] == b['id']]
                assert mine and mine[0]['read_count'] == 1
                audit = await server.db.audit_logs.find_one(
                    {'action': 'admin.broadcast_send', 'target_id': b['id']})
                assert audit, 'broadcast send must be audited'
        run(flow())

    def test_scheduled_broadcast_does_not_send_yet(self, actors):
        h = actors['admin1'][1]
        patient = actors['patient'][0]
        async def flow():
            async with make_client() as cli:
                r = await cli.post('/api/admin/messages', headers=h, json={
                    'subject': f'Maintenance window {RUN_ID}',
                    'body': 'Scheduled maintenance tonight.',
                    'target': f"user:{patient['id']}",
                    'scheduleAt': (server.now() + timedelta(hours=6)).isoformat()})
                assert r.status_code == 200, r.text
                b = r.json()
                assert b['status'] == 'scheduled'
                fanned = await server.db.notifications.find_one({'broadcast_id': b['id']})
                assert fanned is None, 'a scheduled broadcast must not fan out yet'
        run(flow())

    def test_scheduled_broadcast_worker_delivers_exactly_once(self, actors):
        """The background worker claims a due broadcast atomically, fans out
        idempotently, survives a crash mid-fan-out, and never double-sends."""
        import admin_routes
        h = actors['admin1'][1]
        patient = actors['patient'][0]
        async def flow():
            db = server.db
            async with make_client() as cli:
                r = await cli.post('/api/admin/messages', headers=h, json={
                    'subject': f'Worker window {RUN_ID}',
                    'body': 'Scheduled maintenance tonight.',
                    'target': f"user:{patient['id']}",
                    'scheduleAt': (server.now() + timedelta(hours=6)).isoformat()})
                assert r.status_code == 200, r.text
                b = r.json()
                assert b['status'] == 'scheduled'
            # Simulate the schedule coming due.
            await db.broadcast_messages.update_one(
                {'id': b['id']},
                {'$set': {'scheduleAt': server.now() - timedelta(minutes=1)}})
            delivered = await admin_routes.deliver_due_broadcasts(only_id=b["id"])
            assert delivered >= 1
            fresh = await db.broadcast_messages.find_one({'id': b['id']}, {'_id': 0})
            assert fresh['status'] == 'sent' and fresh['sent_at']
            assert await db.notifications.count_documents(
                {'broadcast_id': b['id']}) == 1
            assert await db.notification_deliveries.count_documents(
                {'broadcast_id': b['id']}) == 1
            audit = await db.audit_logs.find_one(
                {'action': 'admin.broadcast_send', 'target_id': b['id']})
            assert audit and audit.get('scheduled') is True
            # Exactly-once: a second worker pass must not double-deliver.
            again = await admin_routes.deliver_due_broadcasts(only_id=b["id"])
            assert again == 0
            assert await db.notifications.count_documents(
                {'broadcast_id': b['id']}) == 1
            # Crash recovery: a stale 'sending' claim with partial fan-out rows
            # is re-claimed and re-delivered without duplicates.
            await db.broadcast_messages.update_one({'id': b['id']}, {'$set': {
                'status': 'sending',
                'claimed_at': server.now() - timedelta(
                    seconds=admin_routes.BROADCAST_CLAIM_STALE_SEC + 60)}})
            await db.notifications.insert_one({
                'id': str(uuid.uuid4()), 'user_id': patient['id'],
                'title': f'Worker window {RUN_ID}', 'body': 'partial row',
                'kind': 'broadcast', 'broadcast_id': b['id'],
                'read': False, 'created_at': server.now()})
            recovered = await admin_routes.deliver_due_broadcasts(only_id=b["id"])
            assert recovered == 1
            fresh2 = await db.broadcast_messages.find_one({'id': b['id']}, {'_id': 0})
            assert fresh2['status'] == 'sent'
            assert await db.notifications.count_documents(
                {'broadcast_id': b['id']}) == 1, 'partial rows must be replaced, not duplicated'
        run(flow())

    def test_replies_respond_and_resolve(self, actors):
        h = actors['admin1'][1]
        patient = actors['patient'][0]
        async def flow():
            db = server.db
            # a sent broadcast + a user reply (replies are filed by the app)
            bid = str(uuid.uuid4())
            await db.broadcast_messages.insert_one({
                'id': bid, 'type': 'general', 'subject': f'Reply thread {RUN_ID}',
                'body': 'b', 'target': 'all', 'allowReplies': True, 'status': 'sent',
                'recipients_count': 1, 'created_by': 'x', 'created_at': server.now(),
                'sent_at': server.now()})
            rid = str(uuid.uuid4())
            await db.broadcast_replies.insert_one({
                'id': rid, 'broadcast_id': bid, 'user_id': patient['id'],
                'user_name': 'A Patient', 'text': f'Question about this {RUN_ID}',
                'status': 'open', 'responses': [], 'created_at': server.now()})
            async with make_client() as cli:
                lst = await cli.get(f'/api/admin/messages/{bid}/replies', headers=h)
                assert lst.status_code == 200 and len(lst.json()) == 1
                resp = await cli.post(f'/api/admin/messages/replies/{rid}/respond',
                                      headers=h, json={'text': f'Answered {RUN_ID}'})
                assert resp.status_code == 200, resp.text
                res = await cli.post(f'/api/admin/messages/replies/{rid}/resolve', headers=h)
                assert res.status_code == 200, res.text
                fresh = await db.broadcast_replies.find_one({'id': rid}, {'_id': 0})
                assert fresh['status'] == 'resolved'
                assert fresh['responses'][0]['text'] == f'Answered {RUN_ID}'
        run(flow())
