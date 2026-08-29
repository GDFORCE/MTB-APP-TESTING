"""Role-scoped report exports (POST /api/reports/generate).

Verifies: role gating, trial-relationship scoping, Sponsor/CRO
de-identification inside the generated file, owner-only recent/download,
and audit records for generation and download.

Same harness as the other modules: in-process ASGITransport, RUN_ID-marked
data, one module-level event loop, module teardown cleanup.
"""
import asyncio
import io
import sys
import uuid
from pathlib import Path

import pytest

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

import httpx  # noqa: E402
import server  # noqa: E402

RUN_ID = uuid.uuid4().hex[:8]
PASSWORD = "Password1!"
ORG = f"ROLE-REPORTS-{RUN_ID} Pharma"
FOREIGN_ORG = f"ROLE-REPORTS-{RUN_ID} Other"
PATIENT_NAME = f"Report Patient {RUN_ID}"
LOOP = asyncio.new_event_loop()

TRIAL_IDS = []
USER_IDS = []
PATIENT_IDS = []
REPORT_IDS = []


def run(coro):
    return LOOP.run_until_complete(coro)


def client():
    return httpx.AsyncClient(
        transport=httpx.ASGITransport(app=server.app),
        base_url="http://testserver",
    )


async def register(role, org):
    email = f"rolerep-{RUN_ID}-{role}-{uuid.uuid4().hex[:5]}@example.com"
    async with client() as cli:
        r = await cli.post("/api/auth/register", json={
            "email": email, "password": PASSWORD,
            "full_name": f"RoleRep {role} {RUN_ID}", "role": role,
            "organization": org,
        })
    assert r.status_code == 200, r.text
    j = r.json()
    USER_IDS.append(j["user"]["id"])
    return j["user"], {"Authorization": f"Bearer {j['access_token']}"}


@pytest.fixture(scope="module")
def world():
    async def build():
        sponsor, sponsor_h = await register("sponsor", ORG)
        foreign, foreign_h = await register("sponsor", FOREIGN_ORG)
        async with client() as cli:
            r = await cli.post("/api/trials", headers=sponsor_h, json={
                "title": f"Role Report Trial {RUN_ID}",
                "protocol_id": f"RR-{RUN_ID}",
                "phase": "Phase II", "condition": "Testing",
                "drug": "Study Drug", "target_enrollment": 20,
            })
        assert r.status_code == 200, r.text
        trial = r.json()
        TRIAL_IDS.append(trial["id"])
        pid = str(uuid.uuid4())
        PATIENT_IDS.append(pid)
        await server.db.patients.insert_one({
            "id": pid, "trial_id": trial["id"], "full_name": PATIENT_NAME,
            "status": "Active", "enrolled_date": "2026-07-01",
            "avatar_initials": "RP", "created_at": server.now(),
        })
        return {"sponsor": sponsor, "sponsor_h": sponsor_h,
                "foreign_h": foreign_h, "trial": trial}
    return run(build())


@pytest.fixture(scope="module", autouse=True)
def cleanup():
    yield
    async def clean():
        db = server.db
        await db.trials.delete_many({"id": {"$in": TRIAL_IDS}})
        await db.users.delete_many({"id": {"$in": USER_IDS}})
        await db.organizations.delete_many({"name": {"$in": [ORG, FOREIGN_ORG]}})
        await db.patients.delete_many({"id": {"$in": PATIENT_IDS}})
        await db.audit_logs.delete_many({"user_id": {"$in": USER_IDS}})
        try:
            import storage as storage_mod
            st = storage_mod.get_storage()
            reports = await db.role_reports.find(
                {"id": {"$in": REPORT_IDS}}, {"_id": 0}).to_list(50)
            for rep in reports:
                await st.delete(rep["key"])
        except Exception:
            pass
        await db.role_reports.delete_many({"id": {"$in": REPORT_IDS}})
    run(clean())
    LOOP.close()


def test_patient_role_cannot_generate(world):
    async def flow():
        patient, patient_h = await register("patient", None)
        async with client() as cli:
            r = await cli.post("/api/reports/generate", headers=patient_h,
                               json={"type": "patient-status", "format": "pdf"})
            assert r.status_code == 403, r.text
    run(flow())


def test_sponsor_report_is_scoped_deidentified_and_audited(world):
    async def flow():
        async with client() as cli:
            r = await cli.post("/api/reports/generate", headers=world["sponsor_h"],
                               json={"type": "patient-status", "format": "xlsx"})
            assert r.status_code == 200, r.text
            rep = r.json()
            REPORT_IDS.append(rep["id"])
            assert rep["deidentified"] is True
            assert rep["rows"] >= 1
            # generation is audited
            audit = await server.db.audit_logs.find_one(
                {"action": "report.generate", "target_id": rep["id"]})
            assert audit and audit.get("user_id") == world["sponsor"]["id"]
            # the exported workbook is scoped + de-identified
            dl = await cli.get(f"/api/reports/{rep['id']}/download",
                               headers=world["sponsor_h"])
            assert dl.status_code == 200, dl.text
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(dl.content))
            cells = " | ".join(
                str(c) for row in wb.active.iter_rows(values_only=True) for c in row if c)
            assert PATIENT_NAME not in cells, "sponsor export must be de-identified"
            assert "SUBJ-" in cells
            assert f"RR-{RUN_ID}" in cells
            # download is audited
            dl_audit = await server.db.audit_logs.find_one(
                {"action": "report.download", "target_id": rep["id"]})
            assert dl_audit and dl_audit.get("user_id") == world["sponsor"]["id"]
    run(flow())


def test_recent_and_download_are_owner_scoped(world):
    async def flow():
        async with client() as cli:
            r = await cli.post("/api/reports/generate", headers=world["sponsor_h"],
                               json={"type": "enrolment-summary", "format": "pdf"})
            assert r.status_code == 200, r.text
            rep = r.json()
            REPORT_IDS.append(rep["id"])
            assert rep["rows"] == 1  # exactly the sponsor's one trial
            # PDF binary signature
            dl = await cli.get(f"/api/reports/{rep['id']}/download",
                               headers=world["sponsor_h"])
            assert dl.status_code == 200 and dl.content[:5] == b"%PDF-"
            # a different account cannot list or download it
            other_recent = await cli.get("/api/reports/recent",
                                         headers=world["foreign_h"])
            assert rep["id"] not in {x["id"] for x in other_recent.json()}
            stolen = await cli.get(f"/api/reports/{rep['id']}/download",
                                   headers=world["foreign_h"])
            assert stolen.status_code == 403, stolen.text
            # a foreign sponsor's scope must not include this trial at all
            foreign_rep = await cli.post(
                "/api/reports/generate", headers=world["foreign_h"],
                json={"type": "enrolment-summary", "format": "pdf"})
            assert foreign_rep.status_code == 200, foreign_rep.text
            REPORT_IDS.append(foreign_rep.json()["id"])
            assert foreign_rep.json()["rows"] == 0
    run(flow())
